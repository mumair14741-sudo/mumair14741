"""
Real User Traffic — unified Real-Traffic + Form-Filler runner.

Each visit = one browser session through a residential proxy with a
UA-parsed device fingerprint. Filters happen BEFORE the click is sent:
    · allowed_os        — parse each UA, skip if OS not allowed
    · allowed_countries — probe proxy exit-IP, skip if country not allowed
    · skip_vpn          — skip if exit-IP is flagged proxy/hosting
    · skip_duplicate_ip — skip if exit-IP already exists in user's clicks
    · no_repeated_proxy — each proxy line used at most once per run

If form_fill_enabled is on, after the tracker click we multi-step-fill the
landing form with a row from the uploaded Excel / Google Sheet, take a
final-page screenshot, capture TrustedForm/LeadID proof, and zip everything.

Output:  results.zip  containing  screenshots/*.png  +  report.xlsx
"""
from __future__ import annotations
import asyncio
import io
import os
import random
import time
import uuid
import zipfile
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

if not os.environ.get("PLAYWRIGHT_BROWSERS_PATH") and os.path.isdir("/pw-browsers"):
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = "/pw-browsers"


# Serialize concurrent "ensure chromium installed" attempts. The preview pod's
# ephemeral filesystem can wipe the chromium binary across restarts, and the
# non-blocking startup hook in server.py may not finish before the first
# RUT job fires — so before each job launch we synchronously verify the
# browser is present, installing it with a lock if missing. This guarantees
# the very first visit NEVER fails with "Executable doesn't exist".
_CHROMIUM_INSTALL_LOCK = asyncio.Lock()
# Tracks whether an install is currently in progress so the engine-status
# API can report "installing" instead of just "missing" while the binary
# is being downloaded.
_CHROMIUM_INSTALL_IN_PROGRESS = False


def get_engine_status() -> Dict[str, Any]:
    """Return the current state of the Playwright chromium-headless-shell
    binary so the frontend can show a coloured "Engine Status" badge:
        ready      → binary present at the EXACT revision Playwright wants
        installing → install_in_progress flag is set
        missing    → binary absent and no install in progress
        error      → couldn't read browsers.json (unexpected)
    """
    browsers_root = os.environ.get("PLAYWRIGHT_BROWSERS_PATH", "/pw-browsers")
    expected: Optional[str] = None
    try:
        import json as _json
        import playwright as _pw
        bj = Path(_pw.__file__).parent / "driver" / "package" / "browsers.json"
        if bj.exists():
            with open(bj, "r") as fh:
                data = _json.load(fh)
            for entry in data.get("browsers", []):
                if entry.get("name") == "chromium-headless-shell":
                    expected = str(entry.get("revision") or "").strip() or None
                    break
    except Exception:
        expected = None

    if not expected:
        return {
            "status": "error",
            "message": "Cannot read Playwright revision metadata",
            "expected_revision": None,
            "browser_path": None,
        }

    binary_path = Path(browsers_root) / f"chromium_headless_shell-{expected}" / "chrome-linux" / "headless_shell"
    if binary_path.exists():
        return {
            "status": "ready",
            "message": f"Chromium rev {expected} ready",
            "expected_revision": expected,
            "browser_path": str(binary_path),
        }
    if _CHROMIUM_INSTALL_IN_PROGRESS:
        return {
            "status": "installing",
            "message": f"Downloading Chromium rev {expected}…",
            "expected_revision": expected,
            "browser_path": str(binary_path),
        }
    return {
        "status": "missing",
        "message": f"Chromium rev {expected} not installed yet",
        "expected_revision": expected,
        "browser_path": str(binary_path),
    }


async def _ensure_chromium_available() -> bool:
    """Returns True when the EXACT chromium-headless-shell revision required
    by the installed Playwright Python package is present (installing it
    first if missing). Safe to call before every job — no-op when binary
    is already present.

    NOTE: Earlier versions of this helper used a glob pattern
    `chromium_headless_shell-*` which matched ANY revision present on disk
    (e.g. an old 1208 left over from a previous Playwright upgrade) and
    falsely returned True even though Playwright 1.49.x specifically wanted
    revision 1148 → BrowserType.launch() blew up with "Executable doesn't
    exist at /pw-browsers/chromium_headless_shell-1148/...". We now read
    the EXACT revision from Playwright's bundled browsers.json and verify
    that specific path exists.
    """
    browsers_root = os.environ.get("PLAYWRIGHT_BROWSERS_PATH", "/pw-browsers")

    def _expected_revision() -> Optional[str]:
        """Read the chromium-headless-shell revision Playwright expects.
        Falls back to None if the JSON layout changes."""
        try:
            import json as _json
            import playwright as _pw
            pw_root = Path(_pw.__file__).parent
            bj = pw_root / "driver" / "package" / "browsers.json"
            if not bj.exists():
                return None
            with open(bj, "r") as fh:
                data = _json.load(fh)
            for entry in data.get("browsers", []):
                if entry.get("name") == "chromium-headless-shell":
                    rev = str(entry.get("revision") or "").strip()
                    return rev or None
        except Exception as e:
            logger.debug(f"_expected_revision: {e}")
        return None

    def _binary_for(rev: Optional[str]) -> Optional[Path]:
        if not rev:
            return None
        return Path(browsers_root) / f"chromium_headless_shell-{rev}" / "chrome-linux" / "headless_shell"

    expected = _expected_revision()

    def _exists() -> bool:
        # Strict check: the EXACT revision Playwright wants must be present.
        if expected:
            bp = _binary_for(expected)
            if bp and bp.exists():
                return True
            return False
        # Fallback (only when we can't read browsers.json): glob check.
        try:
            for p in Path(browsers_root).glob("chromium_headless_shell-*"):
                if (p / "chrome-linux" / "headless_shell").exists():
                    return True
        except Exception:
            pass
        return False

    if _exists():
        return True

    # Missing — install with a lock to prevent duplicate installs when
    # multiple jobs start in parallel on a fresh pod.
    async with _CHROMIUM_INSTALL_LOCK:
        # Re-check after acquiring lock (another coroutine may have just
        # finished the install while we waited).
        if _exists():
            return True
        global _CHROMIUM_INSTALL_IN_PROGRESS
        _CHROMIUM_INSTALL_IN_PROGRESS = True
        try:
            logger.warning(
                f"Playwright chromium-headless-shell rev {expected or '?'} missing — "
                f"installing now (this may take ~60s)…"
            )
            try:
                proc = await asyncio.create_subprocess_exec(
                    "playwright", "install", "chromium-headless-shell",
                    env={**os.environ, "PLAYWRIGHT_BROWSERS_PATH": browsers_root},
                    stdout=asyncio.subprocess.PIPE,
                    stderr=asyncio.subprocess.PIPE,
                )
                try:
                    _out, err = await asyncio.wait_for(proc.communicate(), timeout=300)
                except asyncio.TimeoutError:
                    try: proc.kill()
                    except Exception: pass
                    logger.error("Playwright install timed out after 5 min")
                    return False
                if proc.returncode != 0:
                    logger.error(
                        f"Playwright install returned {proc.returncode}: "
                        f"{(err or b'').decode(errors='ignore')[:300]}"
                    )
                    return False
                logger.info(
                    f"Playwright chromium-headless-shell install: OK (rev {expected or '?'})"
                )
            except Exception as e:
                logger.error(f"Playwright install failed: {e}")
                return False
            # Final strict check — must satisfy the EXACT revision Playwright wants
            return _exists()
        finally:
            _CHROMIUM_INSTALL_IN_PROGRESS = False


import httpx
import pandas as pd
from user_agents import parse as ua_parse
from playwright.async_api import async_playwright, Page, BrowserContext, Browser

from form_filler import (
    load_rows_from_excel,
    load_rows_from_google_sheet,
    _page_has_captcha,
    _dismiss_popups,
    _ensure_form_visible,
    _fill_form,
    _click_submit,
    _dismiss_review_modal,
    _tick_consent_checkboxes,
)

logger = logging.getLogger(__name__)

RESULTS_ROOT = Path("/app/backend/real_user_traffic_results")
RESULTS_ROOT.mkdir(parents=True, exist_ok=True)

RUT_JOBS: Dict[str, Dict[str, Any]] = {}


def _device_name_from_ua(ua_str: str) -> str:
    """Extract a human-readable device name from a user-agent string.

    Strategy:
      1. Let the `user_agents` library try first (handles most Android device
         codes like "SM-S918U" → "Samsung SM-S918U" and "Pixel 8 Pro").
      2. For iOS UAs (which never contain the exact model), fall back to a
         regex over CPU / hardware hints, else return "iPhone" / "iPad".
      3. Windows/macOS/Linux: return OS name + short hardware hint.
    """
    import re
    s = ua_str or ""
    try:
        ua = ua_parse(s)
    except Exception:
        ua = None

    # 1. Library-parsed brand + model (works great for Android)
    if ua and ua.device:
        brand = (ua.device.brand or "").strip()
        model = (ua.device.model or "").strip()
        family = (ua.device.family or "").strip()
        if brand and model and brand.lower() not in model.lower():
            return f"{brand} {model}".strip()
        if model and model.lower() not in ("generic", "smartphone", "other"):
            return model
        if family and family.lower() not in ("generic", "smartphone", "other"):
            return family

    sl = s.lower()

    # 2. iOS — pull iPhone/iPad + model hint if we can, otherwise fall back
    if "iphone" in sl or ("cpu iphone os" in sl):
        # Try to find an explicit model (rare but possible): "iPhone 15 Pro"
        m = re.search(r"iPhone\s*(?:OS\s*)?(\d{1,2}(?:[._]\d{1,2})?)?\s*(?:Max|Pro|Plus|mini)?",
                      s, flags=re.I)
        # Prefer an iOS version tag ("iOS 17.1") to differentiate
        iosv = ""
        m2 = re.search(r"iphone os (\d+[._]\d+)", sl)
        if m2:
            iosv = m2.group(1).replace("_", ".")
        return f"iPhone (iOS {iosv})" if iosv else "iPhone"
    if "ipad" in sl:
        return "iPad"

    # 3. Android fallback when the library couldn't detect a model
    if "android" in sl:
        # Pull the "build" code inside parentheses, e.g. "(Linux; Android 14; SM-S918U)"
        m = re.search(r"android\s+[\d.]+;\s*([^;)\s][^;)]*?)(?:\)|;|\s+build)", s, flags=re.I)
        if m:
            model = m.group(1).strip()
            if model and not model.lower().startswith("linux"):
                return f"Android — {model}"
        return "Android"

    # 4. Desktop
    if "windows" in sl:
        m = re.search(r"windows nt (\d+\.\d+)", sl)
        return f"Windows {m.group(1)}" if m else "Windows PC"
    if "macintosh" in sl or "mac os x" in sl:
        return "Mac"
    if "linux" in sl:
        return "Linux"
    return "Unknown"


# ─── UA → device fingerprint ─────────────────────────────────────
def _os_key_from_ua(ua_str: str) -> str:
    """Return a lowercase OS key matching the frontend's allowed_os chips."""
    try:
        ua = ua_parse(ua_str or "")
        fam = (ua.os.family or "").lower()
    except Exception:
        fam = ""
    if "android" in fam:
        return "android"
    if "ios" in fam or "iphone" in fam or "ipad" in fam:
        return "ios"
    if "windows" in fam:
        return "windows"
    if "mac" in fam:
        return "macos"
    if any(k in fam for k in ("linux", "ubuntu", "fedora", "debian", "chromium os", "chrome os")):
        return "linux"
    return "other"


def _fingerprint_from_ua(ua_str: str) -> Dict[str, Any]:
    """Derive viewport / DPR / platform / mobile flags from a user-agent."""
    try:
        ua = ua_parse(ua_str or "")
    except Exception:
        ua = None

    os_key = _os_key_from_ua(ua_str)
    is_mobile = bool(ua and ua.is_mobile)
    is_tablet = bool(ua and ua.is_tablet)

    # Plausible ranges per OS — picked FRESH each visit for true uniqueness.
    if os_key == "ios":
        platform = "iPhone" if not is_tablet else "iPad"
        vendor = "Apple Computer, Inc."
        viewport = {"width": 390, "height": 844} if not is_tablet else {"width": 820, "height": 1180}
        dpr = 3 if not is_tablet else 2
        hc = random.choice([4, 6])
        dm = random.choice([4, 6, 8])
        webgl_vendor = "Apple Inc."
        webgl_renderer = random.choice([
            "Apple GPU", "Apple A15 GPU", "Apple A16 GPU", "Apple A17 Pro GPU",
        ])
    elif os_key == "android":
        platform = "Linux armv8l"
        vendor = "Google Inc."
        viewport = {"width": 412, "height": 915} if not is_tablet else {"width": 800, "height": 1280}
        dpr = random.choice([2.0, 2.625, 3.0])
        hc = random.choice([6, 8])
        dm = random.choice([4, 6, 8])
        webgl_vendor = "Google Inc. (Qualcomm)"
        webgl_renderer = random.choice([
            "ANGLE (Qualcomm, Adreno (TM) 740, OpenGL ES 3.2)",
            "ANGLE (Qualcomm, Adreno (TM) 730, OpenGL ES 3.2)",
            "ANGLE (ARM, Mali-G78 MP24, OpenGL ES 3.2)",
            "ANGLE (Qualcomm, Adreno (TM) 650, OpenGL ES 3.2)",
        ])
    elif os_key == "windows":
        platform = "Win32"
        vendor = "Google Inc."
        viewport = random.choice([
            {"width": 1920, "height": 1080},
            {"width": 1536, "height": 864},
            {"width": 1366, "height": 768},
            {"width": 1680, "height": 1050},
        ])
        dpr = random.choice([1.0, 1.25, 1.5])
        hc = random.choice([4, 8, 12, 16])
        dm = random.choice([4, 8, 16, 32])
        webgl_vendor = "Google Inc. (NVIDIA)"
        webgl_renderer = random.choice([
            "ANGLE (NVIDIA, NVIDIA GeForce GTX 1650 Direct3D11 vs_5_0 ps_5_0)",
            "ANGLE (NVIDIA, NVIDIA GeForce RTX 3060 Direct3D11 vs_5_0 ps_5_0)",
            "ANGLE (Intel, Intel(R) UHD Graphics 630 Direct3D11 vs_5_0 ps_5_0)",
            "ANGLE (AMD, AMD Radeon RX 6600 Direct3D11 vs_5_0 ps_5_0)",
        ])
    elif os_key == "macos":
        platform = "MacIntel"
        vendor = "Google Inc."
        viewport = random.choice([
            {"width": 1440, "height": 900},
            {"width": 1512, "height": 982},
            {"width": 1680, "height": 1050},
        ])
        dpr = 2
        hc = random.choice([8, 10, 12])
        dm = random.choice([8, 16])
        webgl_vendor = "Google Inc. (Apple)"
        webgl_renderer = random.choice([
            "ANGLE (Apple, Apple M1 Pro, OpenGL 4.1)",
            "ANGLE (Apple, Apple M2, OpenGL 4.1)",
            "ANGLE (Apple, Apple M3, OpenGL 4.1)",
        ])
    elif os_key == "linux":
        platform = "Linux x86_64"
        vendor = "Google Inc."
        viewport = {"width": 1920, "height": 1080}
        dpr = 1
        hc = random.choice([4, 8, 12])
        dm = random.choice([4, 8, 16])
        webgl_vendor = "Google Inc. (Intel)"
        webgl_renderer = "ANGLE (Intel, Mesa Intel(R) UHD Graphics 620)"
    else:
        platform = "Linux x86_64"
        vendor = "Google Inc."
        viewport = {"width": 1366, "height": 768}
        dpr = 1
        hc, dm = 4, 8
        webgl_vendor = "Google Inc."
        webgl_renderer = "ANGLE (Intel, Mesa Intel(R) HD Graphics)"

    # Small jitter on top so even two visits from the same preset look distinct
    viewport = {
        "width": max(320, viewport["width"] + random.randint(-4, 4)),
        "height": max(568, viewport["height"] + random.randint(-8, 8)),
    }

    return {
        "os": os_key,
        "platform": platform,
        "vendor": vendor,
        "viewport": viewport,
        "device_scale_factor": dpr,
        "is_mobile": is_mobile or is_tablet or os_key in ("android", "ios"),
        "has_touch": is_mobile or is_tablet or os_key in ("android", "ios"),
        "hardware_concurrency": hc,
        "device_memory": dm,
        "webgl_vendor": webgl_vendor,
        "webgl_renderer": webgl_renderer,
        # Canvas noise seed — unique per visit so canvas fingerprint differs too
        "canvas_seed": random.randint(1, 2**30),
        "label": f"{(ua.os.family + ' ' + ua.os.version_string) if ua else os_key}".strip() or ua_str[:40],
    }


# ─── Proxy helpers ───────────────────────────────────────────────
def _parse_proxy_line(line: str) -> Optional[Dict[str, Any]]:
    s = (line or "").strip()
    if not s:
        return None
    scheme = "http"
    if s.startswith("http://"):
        s = s[7:]
    elif s.startswith("https://"):
        s = s[8:]
        scheme = "https"
    user, pwd = None, None
    if "@" in s:
        auth, s = s.rsplit("@", 1)
        if ":" in auth:
            user, pwd = auth.split(":", 1)
        else:
            user = auth
    parts = s.split(":")
    if len(parts) == 2:
        host, port = parts
    elif len(parts) == 4:
        host, port, user, pwd = parts
    else:
        return None
    try:
        int(port)
    except ValueError:
        return None
    out: Dict[str, Any] = {"server": f"{scheme}://{host}:{port}", "raw": line.strip()}
    if user:
        out["username"] = user
    if pwd:
        out["password"] = pwd
    return out


async def _probe_proxy_geo(proxy: Dict[str, Any], ua: str) -> Dict[str, Any]:
    """Probe proxy through ip-api — returns exit IP + country + city + timezone +
    locale + accept_language + is_vpn flag."""
    result = {
        "exit_ip": None, "country": "US", "country_name": "United States",
        "city": "New York", "region": "NY", "region_name": "New York",
        "lat": 40.7128, "lon": -74.0060,
        "timezone": "America/New_York", "accept_language": "en-US,en;q=0.9",
        "locale": "en-US", "is_vpn": False, "ok": False,
    }
    server = proxy["server"]
    if proxy.get("username"):
        prefix, rest = server.split("://", 1)
        server = f"{prefix}://{proxy['username']}:{proxy.get('password','')}@{rest}"

    # Some commercial residential proxies (proxy-jet, brightdata, etc.) ONLY accept
    # HTTPS CONNECT tunnels and reject plain `GET http://…` forward-proxy requests,
    # so we try an HTTPS geolocation endpoint first. If that fails we fall back to
    # the original HTTP ip-api.com endpoint (which works on proxies that do allow
    # plain HTTP forwarding).
    async def _try_https_ipwhois(cli: httpx.AsyncClient) -> bool:
        try:
            r = await cli.get("https://ipwho.is/")
            if r.status_code == 200:
                data = r.json()
                if data.get("success") is True:
                    result["exit_ip"] = data.get("ip")
                    result["country_name"] = data.get("country") or result["country_name"]
                    result["country"] = data.get("country_code") or result["country"]
                    result["region_name"] = data.get("region") or result["region_name"]
                    result["region"] = data.get("region_code") or result["region"]
                    result["city"] = data.get("city") or result["city"]
                    try:
                        result["lat"] = float(data.get("latitude") or result["lat"])
                        result["lon"] = float(data.get("longitude") or result["lon"])
                    except (TypeError, ValueError):
                        pass
                    tz = data.get("timezone") or {}
                    if isinstance(tz, dict):
                        result["timezone"] = tz.get("id") or result["timezone"]
                    elif isinstance(tz, str):
                        result["timezone"] = tz or result["timezone"]
                    conn = data.get("connection") or {}
                    result["is_vpn"] = bool(
                        conn.get("type") in ("hosting", "datacenter")
                        or (str(conn.get("org") or "").lower().find("hosting") >= 0)
                    )
                    return True
        except Exception as e:
            logger.debug(f"ipwho.is probe failed: {e}")
        return False

    async def _try_http_ipapi(cli: httpx.AsyncClient) -> bool:
        try:
            r = await cli.get(
                "http://ip-api.com/json/?fields=status,country,countryCode,region,regionName,city,"
                "timezone,lat,lon,query,proxy,hosting"
            )
            if r.status_code == 200:
                data = r.json()
                if data.get("status") == "success":
                    result["exit_ip"] = data.get("query")
                    result["country_name"] = data.get("country") or result["country_name"]
                    result["country"] = data.get("countryCode") or result["country"]
                    result["region"] = data.get("region") or result["region"]
                    result["region_name"] = data.get("regionName") or result["region_name"]
                    result["city"] = data.get("city") or result["city"]
                    result["lat"] = float(data.get("lat") or result["lat"])
                    result["lon"] = float(data.get("lon") or result["lon"])
                    result["timezone"] = data.get("timezone") or result["timezone"]
                    result["is_vpn"] = bool(data.get("proxy") or data.get("hosting"))
                    return True
        except Exception as e:
            logger.debug(f"ip-api.com probe failed: {e}")
        return False

    try:
        # Longer timeout because residential proxies can take 10-15s to route.
        # Retry up to 3 times — residential proxies (proxy-jet, brightdata,
        # etc.) have ~10-20% per-request failure rate due to rotating exit
        # nodes; retrying the same proxy usually succeeds with a different
        # exit IP on the next attempt.
        timeout_cfg = httpx.Timeout(30.0, connect=20.0)
        ok = False
        for attempt in range(3):
            try:
                async with httpx.AsyncClient(proxy=server, timeout=timeout_cfg, headers={"User-Agent": ua}, verify=False, http2=False) as cli:
                    ok = await _try_https_ipwhois(cli)
                    if not ok:
                        ok = await _try_http_ipapi(cli)
                if ok:
                    break
            except Exception as e:
                logger.debug(f"Proxy probe attempt {attempt+1} failed: {e}")
            # Brief backoff before next attempt
            if attempt < 2:
                await asyncio.sleep(1.5 * (attempt + 1))
        if ok:
            cc = (result["country"] or "").lower()
            lang_map = {
                "us": "en-US,en;q=0.9", "gb": "en-GB,en;q=0.9", "ca": "en-CA,en;q=0.9",
                "au": "en-AU,en;q=0.9", "nz": "en-NZ,en;q=0.9",
                "de": "de-DE,de;q=0.9,en;q=0.7", "fr": "fr-FR,fr;q=0.9,en;q=0.7",
                "es": "es-ES,es;q=0.9,en;q=0.7", "it": "it-IT,it;q=0.9,en;q=0.7",
                "nl": "nl-NL,nl;q=0.9,en;q=0.7", "pt": "pt-PT,pt;q=0.9,en;q=0.7",
                "br": "pt-BR,pt;q=0.9,en;q=0.7", "mx": "es-MX,es;q=0.9,en;q=0.7",
                "jp": "ja-JP,ja;q=0.9,en;q=0.7", "kr": "ko-KR,ko;q=0.9,en;q=0.7",
                "in": "en-IN,en;q=0.9,hi;q=0.8", "pk": "en-PK,en;q=0.9,ur;q=0.8",
                "ae": "ar-AE,ar;q=0.9,en;q=0.8", "sa": "ar-SA,ar;q=0.9,en;q=0.8",
            }
            locale_map = {
                "us": "en-US", "gb": "en-GB", "ca": "en-CA", "au": "en-AU", "nz": "en-NZ",
                "de": "de-DE", "fr": "fr-FR", "es": "es-ES", "it": "it-IT", "nl": "nl-NL",
                "pt": "pt-PT", "br": "pt-BR", "mx": "es-MX", "jp": "ja-JP", "kr": "ko-KR",
                "in": "en-IN", "pk": "en-PK", "ae": "ar-AE", "sa": "ar-SA",
            }
            result["accept_language"] = lang_map.get(cc, "en-US,en;q=0.9")
            result["locale"] = locale_map.get(cc, "en-US")
            result["ok"] = True
    except Exception as e:
        logger.debug(f"Proxy geo probe failed: {e}")
    return result


# ─── Stealth init script ────────────────────────────────────────
def _build_stealth_script(fp: Dict[str, Any], geo: Dict[str, Any]) -> str:
    langs = [s.split(";")[0].strip() for s in geo["accept_language"].split(",") if s.strip()]
    langs = [lg for lg in langs if lg]
    lang_json = "[" + ",".join(f'"{lg}"' for lg in langs[:4]) + "]"
    return f"""
(() => {{
  try {{ Object.defineProperty(navigator, 'webdriver', {{ get: () => false }}); }} catch(e) {{}}
  try {{ if (!window.chrome) window.chrome = {{}}; if (!window.chrome.runtime) window.chrome.runtime = {{}}; }} catch(e) {{}}
  try {{ Object.defineProperty(navigator, 'platform', {{ get: () => {fp['platform']!r} }}); }} catch(e) {{}}
  try {{ Object.defineProperty(navigator, 'vendor',   {{ get: () => {fp['vendor']!r}   }}); }} catch(e) {{}}
  try {{ Object.defineProperty(navigator, 'hardwareConcurrency', {{ get: () => {int(fp['hardware_concurrency'])} }}); }} catch(e) {{}}
  try {{ Object.defineProperty(navigator, 'deviceMemory',        {{ get: () => {int(fp['device_memory'])} }}); }} catch(e) {{}}
  try {{ Object.defineProperty(navigator, 'languages',           {{ get: () => {lang_json} }}); }} catch(e) {{}}
  try {{
    const orig = navigator.permissions.query.bind(navigator.permissions);
    navigator.permissions.query = (p) => p && p.name === 'notifications'
      ? Promise.resolve({{ state: Notification.permission, onchange: null }}) : orig(p);
  }} catch(e) {{}}
  try {{
    if (navigator.plugins.length === 0) {{
      Object.defineProperty(navigator, 'plugins', {{ get: () => [
        {{ name: 'PDF Viewer', filename: 'internal-pdf-viewer' }},
        {{ name: 'Chrome PDF Viewer', filename: 'internal-pdf-viewer' }},
        {{ name: 'Chromium PDF Viewer', filename: 'internal-pdf-viewer' }}
      ]}});
    }}
  }} catch(e) {{}}
  try {{
    const getParameter = WebGLRenderingContext.prototype.getParameter;
    WebGLRenderingContext.prototype.getParameter = function(param) {{
      if (param === 37445) return {fp['webgl_vendor']!r};       // UNMASKED_VENDOR_WEBGL
      if (param === 37446) return {fp['webgl_renderer']!r};     // UNMASKED_RENDERER_WEBGL
      return getParameter.call(this, param);
    }};
    // Also cover WebGL2
    if (window.WebGL2RenderingContext) {{
      const getParameter2 = WebGL2RenderingContext.prototype.getParameter;
      WebGL2RenderingContext.prototype.getParameter = function(param) {{
        if (param === 37445) return {fp['webgl_vendor']!r};
        if (param === 37446) return {fp['webgl_renderer']!r};
        return getParameter2.call(this, param);
      }};
    }}
  }} catch(e) {{}}
  // Canvas fingerprint noise — unique per visit via seed
  try {{
    const SEED = {int(fp['canvas_seed'])};
    // tiny deterministic PRNG seeded per visit
    let _rng_s = SEED;
    const rng = () => {{ _rng_s = (_rng_s * 1664525 + 1013904223) >>> 0; return _rng_s; }};
    const origToDataURL = HTMLCanvasElement.prototype.toDataURL;
    HTMLCanvasElement.prototype.toDataURL = function(...args) {{
      const ctx = this.getContext('2d');
      if (ctx) {{ try {{
        const w = this.width, h = this.height;
        if (w > 0 && h > 0 && w * h < 2000000) {{
          const data = ctx.getImageData(0, 0, w, h);
          for (let i = 0; i < data.data.length; i += 4) {{
            const r = rng() & 3;
            data.data[i]   ^= (r & 1);
            data.data[i+1] ^= (r >> 1) & 1;
          }}
          ctx.putImageData(data, 0, 0);
        }}
      }} catch(e) {{}} }}
      return origToDataURL.apply(this, args);
    }};
    // Also spoof getImageData for readback-style fingerprinting
    const origGetImageData = CanvasRenderingContext2D.prototype.getImageData;
    CanvasRenderingContext2D.prototype.getImageData = function(...args) {{
      const d = origGetImageData.apply(this, args);
      if (d && d.data) {{
        for (let i = 0; i < d.data.length; i += 4) {{
          const r = rng() & 1;
          d.data[i+3] ^= r; // alpha jitter
        }}
      }}
      return d;
    }};
  }} catch(e) {{}}
}})();
"""


# ─── Validation-error detection (invalid data on landing) ──────────
# After a form submit, we scan the page for classic server-side / inline
# validation errors. DISABLED BY DEFAULT — too many landing pages show
# consent / marketing banners that use `.alert-danger` / `.text-danger`
# classes and give false positives on the form page itself. User opts in
# via `invalid_detection_enabled=true` on the create-job call.
# Selectors are the TIGHT set — only field-level validation.
_VALIDATION_ERROR_SELECTORS = [
    ".invalid-feedback",                       # Bootstrap per-field
    ".field-error",                            # common custom class
    ".form-field-error",
    ".ng-invalid-message",                     # Angular Material
    ".Mui-error + .MuiFormHelperText-root",    # MUI
    "input.is-invalid + .invalid-feedback",
    "[aria-invalid='true'] + .error",
    "[aria-invalid='true'] + .form-error",
]

_VALIDATION_ERROR_PHRASES = [
    # Must combine validation verb AND field noun to avoid matching
    # promotional / consent text on the form page itself.
    "invalid email", "invalid e-mail", "invalid zip", "invalid zipcode",
    "invalid postal", "invalid postcode", "invalid phone", "invalid number",
    "invalid address", "invalid date",
    "please enter a valid email", "please enter a valid phone",
    "please enter a valid zip", "please enter a valid address",
    "not a valid email", "not a valid phone", "not a valid zip",
    "enter a valid email address", "enter a valid phone number",
    "enter a valid zip code", "enter a valid postal code",
    # duplicate / already — be specific
    "already registered", "already submitted", "duplicate submission",
    "duplicate email", "already exists in our system", "already in our system",
    # explicit submission failure
    "submission failed", "could not be submitted", "validation failed",
]


# ─── US state matching — map + normaliser ──────────────────────────
# Used by the "match lead state to proxy IP state" feature so a row from
# California only gets submitted via a CA-exit proxy, etc. We accept both
# 2-letter codes ("CA") and full names ("California") on both sides.
_US_STATES = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas",
    "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware",
    "FL": "Florida", "GA": "Georgia", "HI": "Hawaii", "ID": "Idaho",
    "IL": "Illinois", "IN": "Indiana", "IA": "Iowa", "KS": "Kansas",
    "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine", "MD": "Maryland",
    "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota", "MS": "Mississippi",
    "MO": "Missouri", "MT": "Montana", "NE": "Nebraska", "NV": "Nevada",
    "NH": "New Hampshire", "NJ": "New Jersey", "NM": "New Mexico", "NY": "New York",
    "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio", "OK": "Oklahoma",
    "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island", "SC": "South Carolina",
    "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas", "UT": "Utah",
    "VT": "Vermont", "VA": "Virginia", "WA": "Washington", "WV": "West Virginia",
    "WI": "Wisconsin", "WY": "Wyoming", "DC": "District of Columbia",
    "PR": "Puerto Rico", "GU": "Guam", "VI": "U.S. Virgin Islands",
}
_US_NAME_TO_CODE = {v.lower(): k for k, v in _US_STATES.items()}
_US_CODES = set(_US_STATES.keys())


def _normalize_state(s: Any) -> str:
    """Normalise a US state value to its 2-letter uppercase code.
    Accepts codes ("ca", "CA"), full names ("California"), mixed case, with
    trailing whitespace. Returns '' if not recognised."""
    if s is None:
        return ""
    txt = str(s).strip()
    if not txt:
        return ""
    up = txt.upper()
    if up in _US_CODES:
        return up
    low = txt.lower()
    if low in _US_NAME_TO_CODE:
        return _US_NAME_TO_CODE[low]
    # Handle "California, USA" / "California (CA)" / "NJ (New Jersey)" style —
    # try each part separately against codes and names
    import re
    # Split on any separator: comma / paren / slash / dash / pipe
    parts = re.split(r"[,()\/|\-]", txt)
    for raw in parts:
        part = raw.strip()
        if not part:
            continue
        if part.upper() in _US_CODES:
            return part.upper()
        if part.lower() in _US_NAME_TO_CODE:
            return _US_NAME_TO_CODE[part.lower()]
    # 2-letter-only fallback: strip non-alpha → if exactly 2 uppercase letters matching a code
    stripped = re.sub(r"[^A-Za-z]", "", txt).upper()
    if len(stripped) == 2 and stripped in _US_CODES:
        return stripped
    return ""


def _find_state_column(rows: List[Dict[str, Any]]) -> Optional[str]:
    """Return the key in the row dicts that holds the US-state value, or None.
    Looks for common name variations (state, State, region, st, state_code, etc.)."""
    if not rows:
        return None
    # Gather all unique keys across first few rows
    seen = []
    for r in rows[:10]:
        for k in r.keys():
            if k not in seen:
                seen.append(k)
    # Priority order
    priority = [
        "state", "State", "STATE",
        "state_code", "stateCode", "StateCode",
        "region", "Region", "REGION",
        "st", "ST",
        "province", "Province",
    ]
    for p in priority:
        if p in seen:
            return p
    # Fallback: case-insensitive match on any key ending in 'state'
    for k in seen:
        if k.strip().lower() in ("state", "st", "region", "state_code", "statecode"):
            return k
    return None


async def _detect_validation_errors(page: Page) -> Tuple[bool, str]:
    """Scan a page for inline / server-side validation errors.

    Returns (is_invalid, error_message).  Safe — any exception → (False, '').
    """
    # 1. Visible elements with error classes
    try:
        for sel in _VALIDATION_ERROR_SELECTORS:
            try:
                els = await page.query_selector_all(sel)
            except Exception:
                continue
            for el in els:
                try:
                    if not await el.is_visible():
                        continue
                    txt = ((await el.inner_text()) or "").strip()
                except Exception:
                    continue
                if txt and 2 < len(txt) < 400:
                    # Skip benign "required" labels that show up when field is empty
                    # (we only care about validation that FAILED after submit).
                    low = txt.lower()
                    if low in ("required", "*required", "required field"):
                        continue
                    return True, txt[:200]
    except Exception:
        pass

    # 2. Body-text phrase scan (catch server-rendered error banners)
    try:
        body_text = await page.evaluate(
            "() => (document.body ? document.body.innerText : '').toLowerCase().slice(0, 8000)"
        )
    except Exception:
        body_text = ""
    if body_text:
        for phrase in _VALIDATION_ERROR_PHRASES:
            if phrase in body_text:
                # grab a nearby snippet for context
                idx = body_text.find(phrase)
                start = max(0, idx - 30)
                end = min(len(body_text), idx + 120)
                snippet = body_text[start:end].strip().replace("\n", " ")
                return True, snippet[:200]

    return False, ""


# ─── Job runner ──────────────────────────────────────────────────
async def run_real_user_traffic_job(
    job_id: str,
    target_url: str,
    proxies_raw: List[str],
    user_agents: List[str],
    total_clicks: int,
    concurrency: int,
    duration_minutes: float,
    allowed_os: List[str],
    allowed_countries_lc: List[str],
    skip_duplicate_ip: bool,
    skip_vpn: bool,
    follow_redirect: bool,
    no_repeated_proxy: bool,
    form_fill_enabled: bool,
    rows: Optional[List[Dict[str, Any]]],
    skip_captcha: bool,
    duplicate_ip_set: Optional[set],
    post_submit_wait: int = 6,
    automation_steps: Optional[List[Dict[str, Any]]] = None,
    self_heal: bool = True,
    state_match_enabled: bool = False,
    target_mode: str = "clicks",                # "clicks" | "conversions"
    target_conversions: int = 0,
    max_attempts: int = 0,
    invalid_detection_enabled: bool = False,    # OFF by default — consent-text
                                                # banners were causing false positives
    db=None,
    link_id: Optional[str] = None,
    link_owner_id: Optional[str] = None,
    link_short_code: Optional[str] = None,
    # Per-use immediate removal of consumed items from the saved
    # "Uploaded Things" batches. As soon as a proxy / UA is picked for a
    # visit (or a row index is successfully submitted), it is pulled
    # from the saved batch in MongoDB / overwritten in the on-disk XLSX.
    # User explicitly asked for this real-time behaviour rather than a
    # batched end-of-job consume.
    engine_user_id: Optional[str] = None,
    upload_proxy_id: Optional[str] = None,
    upload_ua_id: Optional[str] = None,
    upload_data_file_id: Optional[str] = None,
):
    """
    Main orchestrator. Emits progress into RUT_JOBS[job_id].
    """
    # Guarantee chromium is installed BEFORE launching any visits.
    # This is the single robust guard that recovers from pod restarts that
    # wipe ad-hoc browser installs. First job on a fresh pod will pause
    # here for ~30-60s while the install runs; subsequent jobs are no-ops
    # (binary already present).
    try:
        push_live_step(job_id, 0, "preflight", "info", "Verifying browser engine…")
    except Exception:
        pass
    ok = await _ensure_chromium_available()
    if not ok:
        await _finalise_and_persist(db, job_id, "failed",
                  "Playwright chromium-headless-shell could not be installed. "
                  "Please contact support or retry — the install will be attempted again on the next job.")
        return

    parsed_proxies: List[Dict[str, Any]] = []
    for ln in proxies_raw:
        p = _parse_proxy_line(ln)
        if p:
            parsed_proxies.append(p)
    if not parsed_proxies:
        await _finalise_and_persist(db, job_id, "failed", "No valid proxies after parsing")
        return

    uas = [u.strip() for u in user_agents if u and u.strip()]
    if not uas:
        await _finalise_and_persist(db, job_id, "failed", "No user agents provided")
        return

    # Pre-filter UAs by allowed_os
    allowed_os_set = set((allowed_os or []))
    if allowed_os_set:
        uas_ok = [u for u in uas if _os_key_from_ua(u) in allowed_os_set]
        if not uas_ok:
            sample_detect = [(u[:60], _os_key_from_ua(u)) for u in uas[:3]]
            await _finalise_and_persist(job_id=job_id, db=db, status="failed",
                      error=(
                          f"All UAs filtered by allowed_os={sorted(allowed_os_set)}. "
                          f"Detected: {sample_detect}"
                      ))
            return
    else:
        uas_ok = uas

    job_dir = RESULTS_ROOT / job_id
    shots_dir = job_dir / "screenshots"
    shots_dir.mkdir(parents=True, exist_ok=True)

    total = max(1, min(int(total_clicks), 100000))
    delay_between = (duration_minutes * 60.0 / total) if duration_minutes and duration_minutes > 0 else 0.0

    RUT_JOBS[job_id].update({
        "status": "running",
        "total": total,
        "processed": 0,
        "succeeded": 0,
        "conversions": 0,
        "skipped_captcha": 0,
        "skipped_country": 0,
        "skipped_os": 0,
        "skipped_duplicate_ip": 0,
        "skipped_vpn": 0,
        "skipped_state_mismatch": 0,
        "invalid_data": 0,
        "failed": 0,
        "started_at": datetime.now(timezone.utc).isoformat(),
        "events": [],
        "form_fill_enabled": form_fill_enabled,
        "state_match_enabled": state_match_enabled,
        "invalid_detection_enabled": invalid_detection_enabled,
        "target_mode": target_mode if target_mode in ("clicks", "conversions") else "clicks",
        "target_conversions": int(target_conversions or 0) if target_mode == "conversions" else 0,
        "max_attempts": int(max_attempts or 0) if target_mode == "conversions" else 0,
        # Link context — enables RUT visits to be logged as clicks in the
        # link's user DB (mirrors the /api/t/ tracker behaviour so the
        # dashboard Clicks page shows these visits too).
        "link_id": link_id,
        "link_owner_id": link_owner_id,
        "link_short_code": link_short_code,
    })
    if db is not None:
        await _persist(db, job_id)

    # State-matching config — only honour if rows actually have a state column
    state_col: Optional[str] = None
    state_index: Dict[str, List[int]] = {}  # state_code -> list of row indices
    if state_match_enabled and rows:
        state_col = _find_state_column(rows)
        if state_col:
            for idx, r in enumerate(rows):
                code = _normalize_state(r.get(state_col))
                if code:
                    state_index.setdefault(code, []).append(idx)
        if not state_col or not state_index:
            # Turn off the feature quietly if the file doesn't have a state column
            RUT_JOBS[job_id]["state_match_enabled"] = False
            state_match_enabled = False
    RUT_JOBS[job_id]["state_match_column"] = state_col or ""

    # State-match round-robin pointer per state code
    state_rr: Dict[str, int] = {code: 0 for code in state_index}

    # State shared across tasks
    used_proxy_set: set = set()
    used_ua_set: set = set()  # distinct UA strings actually picked for visits
    consumed_row_indices: set = set()   # rows OK-submitted — NOT reused, removed from pending_leads
    invalid_row_indices: set = set()    # rows that triggered a validation error — ALSO removed from pending_leads
    state = {"proxy_idx": 0, "ua_idx": 0, "row_idx": 0, "start_time": time.time()}
    report: List[Dict[str, Any]] = []
    report_lock = asyncio.Lock()

    # ── Cancellation / stop support ─────────────────────────────────
    # Any code path (worker loop, stop endpoint) can set this flag;
    # new visits will short-circuit; in-flight visits finish their current
    # step and exit. Partial results are still zipped.
    cancel_event = asyncio.Event()
    RUT_JOBS[job_id]["_cancel_event"] = cancel_event

    # ── Per-use immediate deletion (real-time pruning) ──────────────
    # User asked: "ek line use hoe wo sath he delete ho jay" — so as soon
    # as a proxy / UA / row gets consumed in a visit we $pull it from the
    # saved upload batch (or rewrite the on-disk XLSX). Fire-and-forget
    # tasks so the visit isn't blocked by Mongo round-trips. We track every
    # task in `_live_pending_tasks` so the job can await all of them
    # before _finalise_and_persist — without this guard, the LAST visit's
    # $pull was reliably lost when the orchestrator finished too quickly
    # (testing agent caught this: consumed_count = N-1 instead of N).
    _live_proxy_pulled: set = set()  # avoid duplicate $pulls
    _live_ua_pulled: set = set()
    _live_pending_tasks: List[asyncio.Task] = []
    _data_file_lock = asyncio.Lock()  # serialise XLSX rewrites
    user_db_truncated = (engine_user_id or "").replace("-", "_")[:20]

    def _spawn_live(coro) -> None:
        """Schedule a live-remove coroutine and remember it so the job can
        await completion at the end. Replaces bare `asyncio.create_task`."""
        try:
            t = asyncio.create_task(coro)
            _live_pending_tasks.append(t)
        except Exception:
            pass

    async def _live_remove_proxy(raw: str):
        if not (engine_user_id and upload_proxy_id and db is not None and raw):
            return
        if raw in _live_proxy_pulled:
            return
        _live_proxy_pulled.add(raw)
        try:
            client = db.client
            user_db = client[f"trackmaster_user_{user_db_truncated}"]
            res = await user_db["uploaded_resources"].update_one(
                {"id": upload_proxy_id, "user_id": engine_user_id, "type": "proxies"},
                {
                    "$pull": {"items": raw},
                    "$set": {"updated_at": datetime.now(timezone.utc).isoformat()},
                    "$inc": {"consumed_count": 1, "item_count": -1},
                },
            )
            # If the batch is now empty, delete it entirely
            if res.modified_count:
                doc = await user_db["uploaded_resources"].find_one(
                    {"id": upload_proxy_id, "user_id": engine_user_id},
                    {"_id": 0, "items": 1},
                )
                if doc and isinstance(doc.get("items"), list) and len(doc["items"]) == 0:
                    await user_db["uploaded_resources"].delete_one(
                        {"id": upload_proxy_id, "user_id": engine_user_id}
                    )
        except Exception as e:
            logger.warning(f"_live_remove_proxy update_one failed: {type(e).__name__}: {e}")

    async def _live_remove_ua(ua: str):
        if not (engine_user_id and upload_ua_id and db is not None and ua):
            return
        if ua in _live_ua_pulled:
            return
        _live_ua_pulled.add(ua)
        try:
            client = db.client
            user_db = client[f"trackmaster_user_{user_db_truncated}"]
            res = await user_db["uploaded_resources"].update_one(
                {"id": upload_ua_id, "user_id": engine_user_id, "type": "user_agents"},
                {
                    "$pull": {"items": ua},
                    "$set": {"updated_at": datetime.now(timezone.utc).isoformat()},
                    "$inc": {"consumed_count": 1, "item_count": -1},
                },
            )
            if res.modified_count:
                doc = await user_db["uploaded_resources"].find_one(
                    {"id": upload_ua_id, "user_id": engine_user_id},
                    {"_id": 0, "items": 1},
                )
                if doc and isinstance(doc.get("items"), list) and len(doc["items"]) == 0:
                    await user_db["uploaded_resources"].delete_one(
                        {"id": upload_ua_id, "user_id": engine_user_id}
                    )
        except Exception as e:
            logger.debug(f"_live_remove_ua failed: {e}")

    async def _live_remove_data_row(row_idx: int):
        """Rewrite the saved data-file XLSX with the consumed/invalid row
        removed. Lock-serialised so concurrent writes don't corrupt the
        file. The on-disk path is read fresh from the upload doc each
        time so a previous flush is always reflected."""
        if not (engine_user_id and upload_data_file_id and db is not None):
            return
        async with _data_file_lock:
            try:
                client = db.client
                user_db = client[f"trackmaster_user_{user_db_truncated}"]
                doc = await user_db["uploaded_resources"].find_one(
                    {"id": upload_data_file_id, "user_id": engine_user_id, "type": "data_file"},
                    {"_id": 0, "file_path": 1, "items": 1},
                )
                if not doc:
                    return
                fp = doc.get("file_path") or ""
                if not fp or not Path(fp).exists():
                    return
                # Load, drop the row, save back. We use openpyxl directly
                # to keep things fast (no pandas roundtrip for 1 row).
                import openpyxl
                wb = openpyxl.load_workbook(fp)
                ws = wb.active
                # row_idx is 0-based against the original-data rows; the
                # XLSX has a header row at row 1, so the actual sheet row
                # is row_idx + 2. After previous deletions the sheet has
                # fewer rows than the original — we therefore work off
                # row VALUES, not indices: scan all data rows and find
                # the one whose original_row_index column matches.
                # Simpler approach: maintain a hidden "_orig_idx" column
                # added on first write so subsequent deletions work
                # against a stable identifier.
                header = [c.value for c in ws[1]] if ws.max_row >= 1 else []
                if "_orig_idx" not in header:
                    # Add the column once, populate with current sheet
                    # row positions (they correspond 1:1 to the source
                    # data file order on first write).
                    col_idx = len(header) + 1
                    ws.cell(row=1, column=col_idx, value="_orig_idx")
                    for r in range(2, ws.max_row + 1):
                        ws.cell(row=r, column=col_idx, value=r - 2)
                    header.append("_orig_idx")
                orig_col = header.index("_orig_idx") + 1
                target_sheet_row = None
                for r in range(2, ws.max_row + 1):
                    val = ws.cell(row=r, column=orig_col).value
                    try:
                        if int(val) == int(row_idx):
                            target_sheet_row = r
                            break
                    except (TypeError, ValueError):
                        continue
                if target_sheet_row:
                    ws.delete_rows(target_sheet_row, 1)
                wb.save(fp)
                wb.close()
                # Update count + bump consumed_count for analytics
                remaining = max(0, (ws.max_row or 1) - 1)
                await user_db["uploaded_resources"].update_one(
                    {"id": upload_data_file_id, "user_id": engine_user_id},
                    {
                        "$set": {
                            "row_count": remaining,
                            "updated_at": datetime.now(timezone.utc).isoformat(),
                        },
                        "$inc": {"consumed_count": 1},
                    },
                )
                # Auto-delete batch + file if completely consumed
                if remaining == 0:
                    try:
                        Path(fp).unlink(missing_ok=True)
                    except Exception:
                        pass
                    await user_db["uploaded_resources"].delete_one(
                        {"id": upload_data_file_id, "user_id": engine_user_id}
                    )
            except Exception as e:
                logger.debug(f"_live_remove_data_row failed: {e}")

    def pick_next_proxy() -> Optional[Dict[str, Any]]:
        """Round-robin pick a proxy, respecting no_repeated_proxy."""
        if no_repeated_proxy:
            for _ in range(len(parsed_proxies)):
                idx = state["proxy_idx"] % len(parsed_proxies)
                state["proxy_idx"] += 1
                raw = parsed_proxies[idx]["raw"]
                if raw not in used_proxy_set:
                    used_proxy_set.add(raw)
                    return parsed_proxies[idx]
            return None
        idx = state["proxy_idx"] % len(parsed_proxies)
        state["proxy_idx"] += 1
        return parsed_proxies[idx]

    def pick_next_ua() -> str:
        idx = state["ua_idx"] % len(uas_ok)
        state["ua_idx"] += 1
        return uas_ok[idx]

    def pick_next_row() -> Optional[Tuple[int, Dict[str, Any]]]:
        """Return (row_index, row_data) — skips rows already consumed (OK-submitted)
        AND rows flagged as invalid_data. If ALL rows are exhausted, returns None
        (caller should stop retrying)."""
        if not rows:
            return None
        total = len(rows)
        # Find a fresh row (not consumed, not invalid)
        for _ in range(total):
            idx = state["row_idx"] % total
            state["row_idx"] += 1
            if idx in consumed_row_indices or idx in invalid_row_indices:
                continue
            return (idx, rows[idx])
        return None  # all rows either used or invalid — nothing left

    def pick_next_row_for_state(state_code: str) -> Optional[Tuple[int, Dict[str, Any]]]:
        """State-matched row picker — returns a fresh row whose state == state_code.
        Round-robin within the candidate set for that state. Skips consumed/invalid.
        Returns None when no eligible row exists for this state."""
        if not state_code or state_code not in state_index:
            return None
        candidates = state_index[state_code]
        if not candidates:
            return None
        total_c = len(candidates)
        start_ptr = state_rr.get(state_code, 0)
        for _ in range(total_c):
            ptr = start_ptr % total_c
            start_ptr += 1
            idx = candidates[ptr]
            if idx in consumed_row_indices or idx in invalid_row_indices:
                continue
            state_rr[state_code] = start_ptr
            return (idx, rows[idx])
        # advance pointer even on failure so next call tries past it
        state_rr[state_code] = start_ptr
        return None

    # ── Resilient shared-browser holder ─────────────────────────────
    # Chromium occasionally crashes mid-job under heavy concurrency or
    # when a misbehaving proxy forces it to tear down. When that happens
    # every subsequent `browser.new_context(...)` throws
    # `TargetClosedError: Target page, context or browser has been closed`
    # and the whole job "fails" with 0 conversions. This holder wraps the
    # shared browser so workers can lazily relaunch it behind an async
    # lock if it ever drops offline.
    _browser_holder: Dict[str, Any] = {"b": None, "pw": None}
    _browser_lock = asyncio.Lock()

    async def _get_live_browser() -> Browser:
        """Return a live Playwright Browser — relaunches on the fly if the
        shared instance crashed / got disconnected. Safe to call from many
        concurrent workers (serialised via _browser_lock)."""
        b = _browser_holder.get("b")
        if b is not None:
            try:
                if b.is_connected():
                    return b
            except Exception:
                pass
        # Browser missing or disconnected — relaunch under lock.
        async with _browser_lock:
            b = _browser_holder.get("b")
            if b is not None:
                try:
                    if b.is_connected():
                        return b
                except Exception:
                    pass
            pw = _browser_holder.get("pw")
            if pw is None:
                # No Playwright runtime yet — create one.
                pw_cm_local = async_playwright()
                pw = await pw_cm_local.__aenter__()
                _browser_holder["pw"] = pw
                _browser_holder["pw_cm"] = pw_cm_local
            logger.warning(
                f"RUT job {job_id}: shared Chromium unavailable — relaunching…"
            )
            try:
                push_live_step(job_id, 0, "engine", "info",
                               "Chromium crashed — relaunching…")
            except Exception:
                pass
            new_b = await pw.chromium.launch(
                headless=True,
                args=[
                    "--no-sandbox",
                    "--disable-dev-shm-usage",
                    "--disable-features=WebRtcHideLocalIpsWithMdns,AutomationControlled",
                    "--disable-blink-features=AutomationControlled",
                    "--force-webrtc-ip-handling-policy=disable_non_proxied_udp",
                ],
            )
            _browser_holder["b"] = new_b
            try:
                push_live_step(job_id, 0, "engine", "ok",
                               "Chromium relaunched — resuming visits")
            except Exception:
                pass
            return new_b

    async def process_one(i: int, shared_browser: Browser):
        # Short-circuit if user pressed Stop — don't waste a proxy/UA on it
        if cancel_event.is_set():
            return

        entry = {
            "visit_index": i + 1,
            "status": "pending",
            "proxy": "",
            "exit_ip": "",
            "country": "",
            "city": "",
            "timezone": "",
            "locale": "",
            "os": "",
            "ua": "",
            "viewport": "",
            "device_name": "",
            "http_status": "",
            "final_url": "",
            "landing_url": "",
            "conversion_page_reached": False,
            "trusted_form": "",
            "lead_id": "",
            "error": "",
            "screenshot": "",
            "timestamp": datetime.now(timezone.utc).isoformat(),
        }

        proxy = pick_next_proxy()
        if not proxy:
            entry["status"] = "failed"
            entry["error"] = "No more proxies available (no_repeated_proxy = on)"
            push_live_step(job_id, i + 1, "setup", "failed", "No proxies available")
            await _record(job_id, entry, report, report_lock, db)
            return

        # Mark this proxy raw as USED (attempted in a visit). Used by the
        # post-job upload-consume hook so only proxies actually consumed
        # in this run get removed from the saved batch — unused proxies
        # remain in the user's "Uploaded Things" library.
        try:
            raw_line = proxy.get("raw") or ""
            if raw_line:
                used_proxy_set.add(raw_line)
                # IMMEDIATE per-use deletion from the saved upload batch
                # (fire-and-forget — don't block the visit).
                _spawn_live(_live_remove_proxy(raw_line))
        except Exception:
            pass

        ua = pick_next_ua()
        # Track UA strings used so the upload-consume hook removes only
        # those that were actually attempted, not the entire UA batch.
        try:
            if ua:
                used_ua_set.add(ua)
                _spawn_live(_live_remove_ua(ua))
        except Exception:
            pass
        fp = _fingerprint_from_ua(ua)

        entry["proxy"] = proxy.get("server", "")
        entry["os"] = fp["os"]
        entry["ua"] = ua
        entry["viewport"] = f"{fp['viewport']['width']}x{fp['viewport']['height']}"
        entry["device_name"] = _device_name_from_ua(ua)
        entry["webgl_renderer"] = fp.get("webgl_renderer", "")
        entry["canvas_seed"] = fp.get("canvas_seed", 0)
        entry["hardware_concurrency"] = fp.get("hardware_concurrency", 0)
        entry["device_memory"] = fp.get("device_memory", 0)
        entry["device_scale_factor"] = fp.get("device_scale_factor", 0)
        push_live_step(job_id, i + 1, "setup", "info",
                       f"Proxy {entry['proxy']} · {entry['device_name']} · {entry['viewport']}")

        # Probe geo (also gives VPN flag)
        geo = await _probe_proxy_geo(proxy, ua)
        entry["exit_ip"] = geo["exit_ip"] or ""
        entry["country"] = geo["country_name"]
        entry["city"] = geo["city"]
        entry["timezone"] = geo["timezone"]
        entry["locale"] = geo["locale"]
        push_live_step(job_id, i + 1, "geo", "ok" if geo["ok"] else "failed",
                       f"Exit {entry['exit_ip'] or '?'} · {entry['country'] or '?'}, {entry['city'] or '?'}")

        if not geo["ok"]:
            entry["status"] = "failed"
            entry["error"] = "Proxy unreachable (ip-api probe failed)"
            return await _record(job_id, entry, report, report_lock, db)

        # Pre-filter: country
        if allowed_countries_lc and geo["country_name"].lower() not in allowed_countries_lc:
            entry["status"] = "skipped_country"
            entry["error"] = f"{geo['country_name']} not in allowed list"
            push_live_step(job_id, i + 1, "filter", "skipped", f"Country not allowed: {geo['country_name']}")
            return await _record(job_id, entry, report, report_lock, db)

        # Pre-filter: VPN
        if skip_vpn and geo["is_vpn"]:
            entry["status"] = "skipped_vpn"
            entry["error"] = "Exit IP is flagged as VPN/hosting"
            push_live_step(job_id, i + 1, "filter", "skipped", "Exit IP flagged as VPN/hosting")
            return await _record(job_id, entry, report, report_lock, db)

        # Pre-filter: duplicate IP
        if skip_duplicate_ip and duplicate_ip_set and geo["exit_ip"] and geo["exit_ip"] in duplicate_ip_set:
            entry["status"] = "skipped_duplicate_ip"
            entry["error"] = "Exit IP already clicked this link before"
            push_live_step(job_id, i + 1, "filter", "skipped", f"Duplicate IP {geo['exit_ip']}")
            return await _record(job_id, entry, report, report_lock, db)

        # Pick form-fill row — state-matched OR sequential
        row_pick = None
        if form_fill_enabled:
            if state_match_enabled and state_col:
                # Match lead state to this proxy's exit-IP state.
                proxy_state_code = _normalize_state(geo.get("region")) or _normalize_state(geo.get("region_name"))
                entry["proxy_state"] = proxy_state_code or ""
                if proxy_state_code:
                    row_pick = pick_next_row_for_state(proxy_state_code)
                if not row_pick:
                    # No lead available for this proxy's state → skip this visit
                    entry["status"] = "skipped_state_mismatch"
                    entry["error"] = (
                        f"No unused lead for state {proxy_state_code or '?'}"
                        if proxy_state_code
                        else "Proxy state unknown (ip-api region missing)"
                    )
                    push_live_step(
                        job_id, i + 1, "filter", "skipped",
                        f"State mismatch: no lead for {proxy_state_code or '?'}",
                    )
                    return await _record(job_id, entry, report, report_lock, db)
            else:
                row_pick = pick_next_row()
        row_index, row = (row_pick if row_pick else (None, None))
        if row is not None:
            entry["row_index"] = (row_index or 0) + 1
            if state_col:
                entry["lead_state"] = _normalize_state(row.get(state_col)) or ""

        browser: Optional[Browser] = None
        context: Optional[BrowserContext] = None
        try:
            # Use the SHARED browser launched once at job start. Per-visit
            # isolation comes from a fresh BrowserContext with its own proxy,
            # cookies, storage, fingerprint, locale, timezone, viewport — which
            # is functionally identical to a fresh browser launch from any
            # detection script's perspective (canvas / WebGL / navigator are
            # all overridden per-context via init script). This drops RAM
            # usage 5-10x vs. per-visit Chromium launches and lets us safely
            # run 15+ concurrent visits without OOM.
            # NOTE: `_get_live_browser()` transparently relaunches Chromium
            # if it has crashed since the last visit — prevents the entire
            # job from failing with `TargetClosedError` the moment one
            # bad proxy or a Chromium bug kills the shared instance.
            browser = await _get_live_browser()
            try:
                context = await browser.new_context(
                    proxy={
                        "server": proxy["server"],
                        **({"username": proxy["username"]} if proxy.get("username") else {}),
                        **({"password": proxy["password"]} if proxy.get("password") else {}),
                    },
                    user_agent=ua,
                    viewport=fp["viewport"],
                    device_scale_factor=fp["device_scale_factor"],
                    is_mobile=fp["is_mobile"],
                    has_touch=fp["has_touch"],
                    locale=geo["locale"],
                    timezone_id=geo["timezone"],
                    geolocation={"latitude": geo["lat"], "longitude": geo["lon"]},
                    permissions=["geolocation"],
                    extra_http_headers={"Accept-Language": geo["accept_language"]},
                )
            except Exception as _nce:
                # new_context can still race with a crash that happened
                # between is_connected() and the call. Give the holder one
                # last chance to relaunch, then retry the context.
                msg = str(_nce)
                if ("closed" in msg.lower()) or ("TargetClosed" in type(_nce).__name__):
                    browser = await _get_live_browser()
                    context = await browser.new_context(
                        proxy={
                            "server": proxy["server"],
                            **({"username": proxy["username"]} if proxy.get("username") else {}),
                            **({"password": proxy["password"]} if proxy.get("password") else {}),
                        },
                        user_agent=ua,
                        viewport=fp["viewport"],
                        device_scale_factor=fp["device_scale_factor"],
                        is_mobile=fp["is_mobile"],
                        has_touch=fp["has_touch"],
                        locale=geo["locale"],
                        timezone_id=geo["timezone"],
                        geolocation={"latitude": geo["lat"], "longitude": geo["lon"]},
                        permissions=["geolocation"],
                        extra_http_headers={"Accept-Language": geo["accept_language"]},
                    )
                else:
                    raise
            await context.add_init_script(_build_stealth_script(fp, geo))

            if True:

                page = await context.new_page()
                push_live_step(job_id, i + 1, "browser", "info", f"Opening {target_url}")
                try:
                    resp = await page.goto(target_url, timeout=45000, wait_until="domcontentloaded")
                    entry["http_status"] = str(resp.status) if resp else ""
                    # Detect chrome-error pages — happens when the residential
                    # proxy's egress tunnel breaks mid-navigation or DNS
                    # fails. Marking these as failures (instead of "ok")
                    # prevents false-positive success counts and triggers
                    # the upstream retry/reporting path correctly.
                    try:
                        cur_url = (page.url or "")
                    except Exception:
                        cur_url = ""
                    if cur_url.startswith("chrome-error://") or cur_url.startswith("chrome://network-error"):
                        entry["status"] = "failed"
                        entry["error"] = f"Browser navigation error (proxy tunnel broken): {cur_url}"
                        push_live_step(job_id, i + 1, "browser", "failed",
                                       f"Navigation error: {cur_url[:80]}")
                        await context.close()
                        return await _record(job_id, entry, report, report_lock, db)
                    # Grab a lightweight landing thumbnail so the Live Activity
                    # modal can prove the browser really loaded the page.
                    try:
                        landing_shot = shots_dir / f"visit_{i+1:05d}_landing.png"
                        await page.screenshot(path=str(landing_shot), full_page=False, timeout=5000)
                        push_live_step(job_id, i + 1, "browser", "ok",
                                       f"Page loaded (HTTP {entry['http_status'] or '?'})",
                                       screenshot=landing_shot.name)
                    except Exception:
                        push_live_step(job_id, i + 1, "browser", "ok",
                                       f"Page loaded (HTTP {entry['http_status'] or '?'})")
                except Exception as e:
                    entry["status"] = "failed"
                    entry["error"] = f"goto failed: {str(e)[:180]}"
                    push_live_step(job_id, i + 1, "browser", "failed", f"goto failed: {str(e)[:100]}")
                    await context.close()
                    return await _record(job_id, entry, report, report_lock, db)

                await page.wait_for_timeout(600 + random.randint(0, 500))

                # Wait for chained redirects (tracker-302 → parent domain →
                # JS-redirect → landing) to fully settle before we inspect
                # the DOM. Without this, on slower offer sites we inspect
                # an intermediate empty page and miss the CTA.
                try:
                    await page.wait_for_load_state("networkidle", timeout=20000)
                except Exception:
                    pass

                # Capture landing URL (post-tracker redirect settle, PRE form fill).
                # Used later to detect "conversion page reached" = host changed
                # after the submit compared to this landing host.
                try:
                    entry["landing_url"] = page.url
                except Exception:
                    pass

                if follow_redirect:
                    # Give the page a bit more time to do any JS redirect
                    try:
                        await page.wait_for_load_state("networkidle", timeout=10000)
                    except Exception:
                        pass

                # No form fill → plain real click, just screenshot
                if not form_fill_enabled:
                    try:
                        entry["final_url"] = page.url
                    except Exception:
                        pass
                    shot_path = shots_dir / f"visit_{i+1:05d}.png"
                    try:
                        await page.screenshot(path=str(shot_path), full_page=False)
                        entry["screenshot"] = shot_path.name
                    except Exception:
                        pass
                    entry["status"] = "ok"
                    push_live_step(job_id, i + 1, "done", "ok",
                                   f"Visit complete → {entry.get('final_url', '')[:120]}",
                                   screenshot=entry.get("screenshot", ""))
                    await context.close()
                    return await _record(job_id, entry, report, report_lock, db)

                # Form-fill path — with validation-error detection + same-session
                # retry using the next available lead row (max 3 invalid retries).
                MAX_INVALID_RETRIES = 3
                retry_attempt = 0
                tried_row_ids: List[int] = []
                if row_index is not None:
                    tried_row_ids.append(row_index)

                while True:
                    if skip_captcha and await _page_has_captcha(page):
                        entry["status"] = "skipped_captcha"
                        entry["error"] = "Captcha detected on landing"
                        push_live_step(job_id, i + 1, "form", "skipped", "Captcha detected — skipping")
                        break

                    push_live_step(
                        job_id, i + 1, "form", "info",
                        (f"Filling form with lead #{entry.get('row_index', '?')}"
                         + (f" (retry {retry_attempt}/{MAX_INVALID_RETRIES})" if retry_attempt else ""))
                        if row else "Filling form",
                    )
                    # If user provided a custom Automation JSON, run that.
                    # Otherwise fall through to the smart auto-fill heuristic.
                    if automation_steps:
                        step_res = await _execute_automation_steps(
                            page, row or {}, automation_steps, skip_captcha=skip_captcha,
                            self_heal=self_heal,
                        )
                    else:
                        # Click through any CTA ("UNLOCK NOW", "Get Started", etc.)
                        # — up to 3 tries because some offers have a 2-step warm-up.
                        await _ensure_form_visible(page, max_tries=3)
                        if skip_captcha and await _page_has_captcha(page):
                            step_res = {"status": "skipped_captcha", "error": "Captcha after CTA click"}
                        else:
                            step_res = await _multi_step_fill(page, row or {})

                    entry["status"] = step_res["status"]
                    if step_res.get("error"):
                        entry["error"] = step_res["error"]

                    # Detect server-side / inline validation errors. Only run
                    # when submit actually went through (ok / submitted-but-no-
                    # redirect); skipped_captcha / no_fields_matched handled above.
                    # GATED by user toggle — default OFF because many landing
                    # pages have consent banners / alerts that trigger false
                    # positives on the form page itself (before submit).
                    is_invalid_data = False
                    val_error = ""
                    if invalid_detection_enabled and entry["status"] in ("submitted_but_no_redirect", "ok"):
                        try:
                            is_invalid_data, val_error = await _detect_validation_errors(page)
                        except Exception:
                            is_invalid_data = False

                    if is_invalid_data and row_index is not None:
                        # Mark the CURRENT row as invalid (drops from pending_leads)
                        async with report_lock:
                            invalid_row_indices.add(row_index)
                        # Per-use immediate deletion from saved data file
                        _spawn_live(_live_remove_data_row(row_index))

                        push_live_step(
                            job_id, i + 1, "submit", "failed",
                            f"Invalid data on lead #{row_index + 1}: {val_error[:110]}",
                        )

                        if retry_attempt < MAX_INVALID_RETRIES:
                            # Pick next available row and reload the form page
                            # to clear previous error state. State-match aware.
                            if state_match_enabled and state_col:
                                proxy_state_code = _normalize_state(geo.get("region")) or _normalize_state(geo.get("region_name"))
                                next_pick = pick_next_row_for_state(proxy_state_code) if proxy_state_code else None
                            else:
                                next_pick = pick_next_row()
                            if not next_pick:
                                entry["status"] = "invalid_data"
                                entry["error"] = f"Invalid: {val_error[:160]} (no more leads to retry)"
                                push_live_step(job_id, i + 1, "form", "failed",
                                               "No more leads available to retry")
                                break

                            row_index, row = next_pick
                            entry["row_index"] = row_index + 1
                            tried_row_ids.append(row_index)
                            retry_attempt += 1

                            # Reload form page so invalid state is cleared
                            try:
                                await page.goto(target_url, timeout=45000,
                                                wait_until="domcontentloaded")
                                await page.wait_for_timeout(700 + random.randint(0, 500))
                                try:
                                    await page.wait_for_load_state("networkidle", timeout=15000)
                                except Exception:
                                    pass
                            except Exception as e:
                                entry["status"] = "invalid_data"
                                entry["error"] = f"Page reload after invalid failed: {str(e)[:120]}"
                                break

                            push_live_step(
                                job_id, i + 1, "form", "info",
                                f"Retry on same form with next lead #{row_index + 1}",
                            )
                            continue  # retry the while loop
                        else:
                            # Max retries reached — finalize as invalid_data
                            entry["status"] = "invalid_data"
                            entry["error"] = f"Invalid (max {MAX_INVALID_RETRIES} retries): {val_error[:140]}"
                            push_live_step(
                                job_id, i + 1, "submit", "failed",
                                f"Max {MAX_INVALID_RETRIES} invalid retries reached",
                            )
                            break

                    # Normal outcome (ok, submitted_but_no_redirect without validation
                    # error, no_fields_matched, skipped_*, etc.) — exit retry loop.
                    break

                # Stash retry metadata
                entry["retry_attempts"] = retry_attempt
                entry["tried_row_indices"] = [r + 1 for r in tried_row_ids]

                push_live_step(
                    job_id, i + 1, "submit",
                    "ok" if entry["status"] == "ok" else ("skipped" if "skipped" in str(entry["status"]) else "failed"),
                    f"{entry['status']}{' — ' + entry['error'] if entry.get('error') else ''}"[:180],
                )

                # Mark the FINAL lead row as CONSUMED if submit succeeded — it
                # will be excluded from the pending_leads.xlsx so the user
                # never reuses it.
                if entry["status"] == "ok" and row_index is not None:
                    async with report_lock:
                        consumed_row_indices.add(row_index)
                    # IMMEDIATE per-use deletion from saved data file
                    _spawn_live(_live_remove_data_row(row_index))

                # Grab TrustedForm / LeadID proofs (if the landing page uses them)
                try:
                    lead = await page.evaluate("""() => {
                        const grab = sel => {
                            const el = document.querySelector(sel);
                            return el ? (el.value || el.getAttribute('value') || '') : '';
                        };
                        return {
                            trusted_form: grab('[name="xxTrustedFormCertUrl"]') || grab('[name="xxTrustedFormToken"]'),
                            lead_id: grab('#leadid_token') || grab('[name="universal_leadid"]') || grab('[name="LeadiD"]'),
                        };
                    }""")
                    entry["trusted_form"] = lead.get("trusted_form", "")
                    entry["lead_id"] = lead.get("lead_id", "")
                except Exception:
                    pass

                try:
                    entry["final_url"] = page.url
                except Exception:
                    pass
                try:
                    await page.wait_for_load_state("networkidle", timeout=5000)
                except Exception:
                    pass

                # Legacy conversion-page signal (host change only) kept in
                # the entry for debugging comparison with the new strict
                # thank-you detection. The strict check runs below after
                # the post-submit wait and overrides conversion_page_reached.
                try:
                    entry["host_changed_after_submit"] = _did_reach_conversion(
                        entry.get("landing_url") or "",
                        entry.get("final_url") or "",
                    )
                except Exception:
                    entry["host_changed_after_submit"] = False

                # Post-submit wait so the "thank-you" / offers-flow page
                # fully renders BEFORE we screenshot it. Default 6s,
                # user-configurable 3–15s.
                if entry["status"] == "ok":
                    wait_ms = max(3000, min(int(post_submit_wait * 1000), 15000))
                    await page.wait_for_timeout(wait_ms)
                else:
                    await page.wait_for_timeout(900)

                # Re-read final_url after the post-submit wait — redirect chains
                # to the actual thank-you page often happen during this sleep.
                try:
                    entry["final_url"] = page.url
                except Exception:
                    pass

                # Detect mid-navigation proxy tunnel breaks — if after the
                # post-submit wait the page is sitting on a Chromium internal
                # error URL, the visit did NOT actually reach any real page.
                # Mark as failed so conversion stats and downstream retries
                # behave correctly.
                _fu = entry.get("final_url", "") or ""
                if _fu.startswith("chrome-error://") or _fu.startswith("chrome://network-error"):
                    entry["status"] = "failed"
                    if not entry.get("error"):
                        entry["error"] = f"Browser navigation error (proxy tunnel broken mid-flow): {_fu}"
                    push_live_step(job_id, i + 1, "nav", "failed",
                                   f"Navigation error after submit: {_fu[:80]}")

                # STRICT thank-you page detection: needs host change + URL
                # keyword + page text keyword. Only TRUE thank-you pages
                # count as conversions and get screenshotted.
                page_title_str = ""
                page_text_str = ""
                try:
                    page_title_str = await page.title()
                except Exception:
                    pass
                try:
                    page_text_str = await page.evaluate(
                        "() => (document.body ? document.body.innerText : '').slice(0, 4000)"
                    )
                except Exception:
                    pass

                try:
                    entry["thank_you_reached"] = _is_thank_you_page(
                        entry.get("landing_url") or "",
                        entry.get("final_url") or "",
                        page_text_str,
                        page_title_str,
                    )
                except Exception:
                    entry["thank_you_reached"] = False

                entry["page_title"] = (page_title_str or "")[:200]
                # Conversion count is now driven STRICTLY by thank-you-reached
                # (not just any host-change). Matches the user's explicit ask:
                # "jo form thanks page tak complete ho wahi conversion count ho".
                entry["conversion_page_reached"] = bool(entry.get("thank_you_reached"))

                # Screenshot logic: ONLY capture the final screenshot when
                # the thank-you page was reached. Drops disk usage and matches
                # the user's ask ("jo form thanks page tak complete ho os ka
                # screenshot ho"). We keep the small landing thumbnail taken
                # earlier for debugging / proof that the browser opened at all.
                if entry.get("thank_you_reached"):
                    shot_path = shots_dir / f"visit_{i+1:05d}_thankyou.png"
                    try:
                        await page.screenshot(path=str(shot_path), full_page=True)
                        entry["screenshot"] = shot_path.name
                    except Exception as e:
                        logger.debug(f"thank-you screenshot failed: {e}")

                # Final live step — always push, include screenshot if we have one
                push_live_step(
                    job_id, i + 1, "done",
                    "ok" if entry["status"] == "ok" else ("skipped" if "skipped" in str(entry["status"]) else "failed"),
                    f"Visit {entry['status']}{' — ✓ converted' if entry.get('thank_you_reached') else ''} → {entry.get('final_url', '')[:120]}",
                    screenshot=entry.get("screenshot", ""),
                )

                # If we are in target-conversions mode and hit the target,
                # flip the cancel flag so the dispatcher stops spawning
                # new visits and in-flight ones wrap up gracefully.
                try:
                    tgt = int(RUT_JOBS[job_id].get("target_conversions") or 0)
                    if tgt > 0 and entry.get("thank_you_reached"):
                        cur = int(RUT_JOBS[job_id].get("conversions") or 0)
                        # +1 because _record hasn't incremented yet for this entry
                        if (cur + 1) >= tgt:
                            cancel_event.set()
                            RUT_JOBS[job_id]["target_reached"] = True
                except Exception:
                    pass

                await context.close()
        except Exception as e:
            entry["status"] = "failed"
            entry["error"] = f"{type(e).__name__}: {str(e)[:180]}"
        finally:
            # Browser is shared across visits — we only close the per-visit
            # context here. The shared browser is closed by the parent job
            # once ALL workers have finished.
            if context is not None:
                try:
                    await context.close()
                except Exception:
                    pass

        await _record(job_id, entry, report, report_lock, db)

    # Launch with concurrency + optional pacing
    semaphore = asyncio.Semaphore(max(1, min(int(concurrency or 1), 20)))
    conc = max(1, min(int(concurrency or 1), 20))

    async def worker(i: int, shared_browser: Browser):
        # Per-visit pacing: target time for this visit = i * delay_between
        if delay_between > 0:
            target_t = state["start_time"] + i * delay_between
            # sleep in small chunks so cancel is responsive
            while time.time() < target_t:
                if cancel_event.is_set():
                    return
                await asyncio.sleep(min(0.5, target_t - time.time()))
        if cancel_event.is_set():
            return
        async with semaphore:
            if cancel_event.is_set():
                return
            await process_one(i, shared_browser)

    # ── Launch ONE shared Chromium browser for the WHOLE job ─────────
    # All visits create their own isolated BrowserContext from this single
    # browser. This is the standard anti-detection pattern (used by
    # Multilogin/GoLogin/AdsPower under the hood) — every context has its
    # own cookies, storage, proxy, fingerprint and is undetectable from
    # the website's side. RAM cost drops 5-10x vs. per-visit launches,
    # which lets concurrency=15 run safely in the pod.
    # NOTE: The browser + Playwright handle are stashed in `_browser_holder`
    # so `_get_live_browser()` can transparently relaunch Chromium if it
    # ever crashes mid-job (prevents the TargetClosedError death-spiral).
    pw_cm = async_playwright()
    pw = await pw_cm.__aenter__()
    _browser_holder["pw"] = pw
    _browser_holder["pw_cm"] = pw_cm
    shared_browser: Optional[Browser] = None
    try:
        shared_browser = await pw.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--disable-features=WebRtcHideLocalIpsWithMdns,AutomationControlled",
                "--disable-blink-features=AutomationControlled",
                "--force-webrtc-ip-handling-policy=disable_non_proxied_udp",
            ],
        )
        _browser_holder["b"] = shared_browser
    except Exception as e:
        try:
            await pw_cm.__aexit__(type(e), e, None)
        except Exception:
            pass
        await _finalise_and_persist(db, job_id, "failed",
                  f"Playwright browser launch failed: {type(e).__name__}: {str(e)[:160]}")
        return
    push_live_step(job_id, 0, "preflight", "ok",
                   f"Shared Chromium ready · concurrency={conc}")

    # ── Dispatcher ──────────────────────────────────────────────────
    # Two modes:
    #   clicks:       run exactly `total` visits (legacy behaviour)
    #   conversions:  keep spawning visits until `target_conversions` hit OR
    #                 `max_attempts` reached. Respects `concurrency`.
    if (
        target_mode == "conversions"
        and int(target_conversions or 0) > 0
    ):
        target_conv = int(target_conversions)
        max_att = int(max_attempts or 0)
        if max_att <= 0:
            max_att = max(target_conv * 20, target_conv + 50)  # safety default
        RUT_JOBS[job_id]["max_attempts"] = max_att
        RUT_JOBS[job_id]["total"] = max_att  # UI progress bar denominator

        attempt_counter = 0
        in_flight: set = set()
        try:
            while True:
                if cancel_event.is_set():
                    break
                cur_conv = int(RUT_JOBS[job_id].get("conversions") or 0)
                if cur_conv >= target_conv:
                    cancel_event.set()
                    RUT_JOBS[job_id]["target_reached"] = True
                    push_live_step(
                        job_id, 0, "done", "ok",
                        f"Target {target_conv} conversions reached — stopping",
                    )
                    break
                if attempt_counter >= max_att:
                    push_live_step(
                        job_id, 0, "done", "info",
                        f"Max {max_att} attempts exhausted — stopping (conversions: {cur_conv}/{target_conv})",
                    )
                    break

                # Fill the pool up to `concurrency` in-flight visits
                while (
                    len(in_flight) < conc
                    and attempt_counter < max_att
                    and not cancel_event.is_set()
                ):
                    t = asyncio.create_task(process_one(attempt_counter, shared_browser))
                    in_flight.add(t)
                    t.add_done_callback(in_flight.discard)
                    attempt_counter += 1

                if not in_flight:
                    await asyncio.sleep(0.2)
                    continue

                # Wait for any visit to finish before re-evaluating target
                await asyncio.wait(
                    in_flight,
                    return_when=asyncio.FIRST_COMPLETED,
                    timeout=1.5,
                )
        except Exception as e:
            logger.warning(f"RUT conversions-mode dispatcher error: {e}")

        # Graceful finish: let in-flight visits complete (they'll respect
        # cancel_event and short-circuit early if it was set)
        if in_flight:
            await asyncio.gather(*in_flight, return_exceptions=True)

        # Update total to reflect actual attempts launched
        RUT_JOBS[job_id]["total"] = attempt_counter
    else:
        tasks = [asyncio.create_task(worker(i, shared_browser)) for i in range(total)]
        try:
            await asyncio.gather(*tasks, return_exceptions=True)
        except Exception as e:
            logger.warning(f"RUT gather error: {e}")

    # ── Close the shared browser & playwright runtime ────────────────
    # Prefer the holder's current browser (may have been relaunched mid-job)
    # over the original `shared_browser` variable.
    final_browser = _browser_holder.get("b") or shared_browser
    try:
        if final_browser is not None:
            await final_browser.close()
    except Exception as e:
        logger.debug(f"shared browser close failed: {e}")
    # Use holder's pw_cm if present (in case _get_live_browser created a
    # new one after the original pw_cm was discarded).
    final_pw_cm = _browser_holder.get("pw_cm") or pw_cm
    try:
        await final_pw_cm.__aexit__(None, None, None)
    except Exception as e:
        logger.debug(f"playwright runtime exit failed: {e}")
    _browser_holder["b"] = None
    _browser_holder["pw"] = None
    _browser_holder["pw_cm"] = None

    # Remember whether Stop was pressed — used later to set final status
    was_cancelled = cancel_event.is_set()

    # Build Excel report + leftover leads + ZIP
    try:
        _write_excel_report(job_dir / "report.xlsx", report)
    except Exception as e:
        logger.warning(f"Excel report failed: {e}")

    # Write leads_with_status.xlsx — ORIGINAL schema + one extra "status" column
    # marking each row as "used" (submit OK), "invalid" (server-side validation
    # rejected), or "not_used". Color-coded:
    #   green  = used        (remove from pending)
    #   red    = invalid     (remove from pending)
    #   orange = not_used    (keeps — goes into pending_leads.xlsx)
    status_path = None
    pending_path = None
    if rows:
        # Build union of column keys across ALL rows (not just rows[0]) so
        # sparse Excel uploads don't silently lose columns.
        seen = set()
        orig_cols: List[str] = []
        for r in rows:
            for k in r.keys():
                if k not in seen:
                    seen.add(k)
                    orig_cols.append(k)

        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font
            from openpyxl.utils import get_column_letter

            status_path = job_dir / "leads_with_status.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Leads"
            headers = orig_cols + ["status"]
            ws.append(headers)
            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="374151")
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
            # Fills
            used_fill = PatternFill("solid", fgColor="C6EFCE")      # light green
            invalid_fill = PatternFill("solid", fgColor="FFC7CE")   # light red
            unused_fill = PatternFill("solid", fgColor="FFE699")    # light orange
            for idx, r in enumerate(rows):
                if idx in consumed_row_indices:
                    status_val, fill = "used", used_fill
                elif idx in invalid_row_indices:
                    status_val, fill = "invalid", invalid_fill
                else:
                    status_val, fill = "not_used", unused_fill
                row_vals = [r.get(c, "") for c in orig_cols] + [status_val]
                ws.append(row_vals)
                excel_row = idx + 2  # +1 header, +1 one-indexed
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=excel_row, column=col_idx).fill = fill
            # Auto-size columns (approximate)
            for col_idx, col_name in enumerate(headers, start=1):
                max_len = max(
                    [len(str(col_name))]
                    + [len(str(r.get(col_name, ""))) for r in rows[:50]]
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)
            wb.save(status_path)
        except Exception as e:
            logger.warning(f"leads_with_status.xlsx write failed: {e}")
            status_path = None

        # ── pending_leads.xlsx — ONLY the unused rows, ready for next run ──
        # Identical schema to the uploaded lead file (no extra columns), so the
        # user can re-upload this file directly as the next run's data source.
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter

            pending_path = job_dir / "pending_leads.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Pending Leads"
            ws.append(orig_cols)
            # Header styling
            try:
                from openpyxl.styles import PatternFill, Font
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill("solid", fgColor="374151")
                for col_idx in range(1, len(orig_cols) + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
            except Exception:
                pass

            pending_count = 0
            for idx, r in enumerate(rows):
                if idx in consumed_row_indices or idx in invalid_row_indices:
                    continue
                ws.append([r.get(c, "") for c in orig_cols])
                pending_count += 1

            for col_idx, col_name in enumerate(orig_cols, start=1):
                max_len = max(
                    [len(str(col_name))]
                    + [len(str(r.get(col_name, ""))) for r in rows[:50]]
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

            wb.save(pending_path)
            RUT_JOBS[job_id]["pending_leads_count"] = pending_count
            RUT_JOBS[job_id]["pending_leads_path"] = str(pending_path)
        except Exception as e:
            logger.warning(f"pending_leads.xlsx write failed: {e}")
            pending_path = None

    zip_path = job_dir / "results.zip"
    try:
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for p in shots_dir.glob("*.png"):
                zf.write(p, arcname=f"screenshots/{p.name}")
            if (job_dir / "report.xlsx").exists():
                zf.write(job_dir / "report.xlsx", arcname="report.xlsx")
            if status_path and status_path.exists():
                zf.write(status_path, arcname="leads_with_status.xlsx")
            if pending_path and pending_path.exists():
                zf.write(pending_path, arcname="pending_leads.xlsx")
    except Exception as e:
        logger.warning(f"zip build failed: {e}")

    # ── Await all pending live-remove tasks BEFORE finalising ────────
    # The fire-and-forget _spawn_live() calls scheduled per-visit $pull /
    # XLSX-rewrite tasks. If we finalise the job before they complete the
    # LAST visit's deletion is sometimes lost (testing agent caught this:
    # consumed_count ended at N-1 instead of N). Drain the queue here so
    # uploaded_resources reflects the FULL set of consumed items by the
    # time the job is marked completed.
    logger.info(
        f"RUT job {job_id}: draining {len(_live_pending_tasks)} live-remove "
        f"tasks (proxy={len(_live_proxy_pulled)} ua={len(_live_ua_pulled)})"
    )
    if _live_pending_tasks:
        try:
            results = await asyncio.wait_for(
                asyncio.gather(*_live_pending_tasks, return_exceptions=True),
                timeout=30.0,
            )
            errs = [r for r in results if isinstance(r, Exception)]
            if errs:
                logger.warning(f"RUT job {job_id}: {len(errs)} live-remove tasks raised: {errs[:3]}")
        except asyncio.TimeoutError:
            logger.warning(
                f"RUT job {job_id}: {len(_live_pending_tasks)} live-remove tasks "
                f"didn't finish in 30s — proceeding with finalise anyway"
            )
        except Exception as e:
            logger.debug(f"live-remove drain error: {e}")

    RUT_JOBS[job_id].update({
        "status": "stopped" if was_cancelled else "completed",
        "finished_at": datetime.now(timezone.utc).isoformat(),
        "report": report[-200:],  # keep last 200 in memory
        "zip_path": str(zip_path),
        "leftover_leads_count": (len(rows) - len(consumed_row_indices) - len(invalid_row_indices)) if rows else 0,
        "consumed_leads_count": len(consumed_row_indices),
        "invalid_leads_count": len(invalid_row_indices),
        # Tracked so the post-finish upload-consume hook can selectively
        # prune ONLY the proxies / UAs actually used during this run from
        # the saved upload batches (not the entire batch).
        "used_proxy_raws": list(used_proxy_set),
        "used_ua_strings": list(used_ua_set),
    })
    # Remove the non-serializable asyncio.Event before any DB persist
    RUT_JOBS[job_id].pop("_cancel_event", None)

    # ── Auto-consume any "Uploaded Things" batches BEFORE persisting ─
    # The live-remove tasks above pull each used proxy / UA / row in
    # real-time. This batched consume is now a SAFETY-NET that mops up
    # anything the live path missed (e.g. a $pull that raced with
    # auto-delete of an empty batch). We run it BEFORE _persist so that
    # by the time the API reports status=completed, the upload doc is
    # already at its final shape — frontend / tests / users will not see
    # a stale snapshot during the brief window between persist and consume.
    consume_upload_ids: List[str] = []
    if db is not None:
        try:
            jr = await db.real_user_traffic_jobs.find_one(
                {"job_id": job_id},
                {"_id": 0, "consume_upload_ids": 1, "user_id": 1, "pending_leads_path": 1},
            )
            if jr:
                consume_upload_ids = jr.get("consume_upload_ids") or []
                uid = jr.get("user_id")
                if consume_upload_ids and uid:
                    try:
                        from server import _consume_uploads
                        await _consume_uploads(
                            uid,
                            consume_upload_ids,
                            used_proxy_raws=list(used_proxy_set),
                            used_ua_strings=list(used_ua_set),
                            pending_leads_path=jr.get("pending_leads_path") or "",
                        )
                        logger.info(
                            f"RUT job {job_id}: pruned {len(consume_upload_ids)} uploaded batch(es) — "
                            f"removed {len(used_proxy_set)} used proxies, "
                            f"{len(used_ua_set)} used UAs (live + safety-net pass)"
                        )
                    except Exception as e:
                        logger.warning(f"RUT job {job_id}: upload consume failed: {e}")
        except Exception as e:
            logger.warning(f"RUT job {job_id}: pre-persist consume hook failed: {e}")

    if db is not None:
        # Mark consume IDs as cleared so we don't double-process later.
        if consume_upload_ids:
            RUT_JOBS[job_id]["consume_upload_ids"] = []
            RUT_JOBS[job_id]["consumed_upload_ids_final"] = consume_upload_ids
        await _persist(db, job_id)


# ──────────────────────────────────────────────────────────────────
# Custom Automation JSON executor
# ──────────────────────────────────────────────────────────────────
# Supported actions: goto, click, fill, select, check, uncheck, press, wait,
# wait_for_selector, wait_for_navigation, scroll, screenshot, evaluate.
# Every step can take: selector, value, ms, timeout, optional, wait_nav.
# Placeholders in `value`:
#   {{row.FIELD}}  or  {{FIELD}}   → Excel row value (case-insensitive)
#   {{random.N}}                    → N-digit random number
#   {{randomletters.N}}             → N random letters
# ──────────────────────────────────────────────────────────────────
def _substitute(template: str, row: Dict[str, Any]) -> str:
    if not isinstance(template, str):
        return template
    import re
    def repl(m):
        key = m.group(1).strip()
        if key.lower().startswith("row."):
            key = key[4:]
        if key.lower().startswith("random."):
            try:
                n = int(key.split(".", 1)[1])
                return "".join(random.choice("0123456789") for _ in range(max(1, n)))
            except Exception:
                return ""
        if key.lower().startswith("randomletters."):
            try:
                n = int(key.split(".", 1)[1])
                import string
                return "".join(random.choice(string.ascii_lowercase) for _ in range(max(1, n)))
            except Exception:
                return ""
        # Case-insensitive row lookup
        for k, v in row.items():
            if str(k).strip().lower() == key.lower():
                return "" if v is None else str(v)
        return ""
    return re.sub(r"\{\{\s*([^}]+?)\s*\}\}", repl, template)


async def _execute_automation_steps(
    page: Page,
    row: Dict[str, Any],
    steps: List[Dict[str, Any]],
    skip_captcha: bool = True,
    self_heal: bool = True,
) -> Dict[str, Any]:
    """Execute a user-provided automation script step-by-step. Returns
    {status, error?, executed_steps}.  Each step format:
        {"action": "click", "selector": "a.btn-primary", "wait_nav": true}
        {"action": "fill",  "selector": "input[name=first]", "value": "{{first}}"}
        {"action": "select","selector": "select[name=dobmonth]", "value": "{{month}}"}
        {"action": "wait",  "ms": 2000}
        {"action": "screenshot", "name": "after_submit"}

    When `self_heal=True`, if a non-optional step fails we take a screenshot
    of the current page and ask Gemini 2.5 Pro for a single recovery action
    (dismiss popup, click Continue, etc.). We try the main step one more time
    after applying the recovery.
    """
    executed = 0
    heal_used = 0
    MAX_HEAL = 3  # total AI recovery attempts per row
    try:
        for idx, step in enumerate(steps or []):
            if not isinstance(step, dict):
                continue
            action = (step.get("action") or "").strip().lower()
            selector = step.get("selector") or ""
            value = _substitute(step.get("value", ""), row)
            timeout = int(step.get("timeout") or 10000)
            optional = bool(step.get("optional") or False)
            wait_nav = bool(step.get("wait_nav") or False)

            if skip_captcha and action not in ("wait", "screenshot", "evaluate"):
                if await _page_has_captcha(page):
                    return {"status": "skipped_captcha", "error": f"Captcha at step {idx+1}", "executed_steps": executed}

            try:
                if action == "goto":
                    await page.goto(value or selector, timeout=timeout, wait_until="domcontentloaded")
                elif action == "click":
                    if wait_nav:
                        # Expect navigation to fire as a result of the click.
                        # Many modern lead-gen pages attach JS handlers to the
                        # submit button that fire analytics/tracking but DO NOT
                        # actually submit the form — so a bare `page.click` +
                        # wait_for_load_state("networkidle") misses the fact
                        # that the form never POSTed. We use expect_navigation
                        # to detect this explicitly, and fall back to calling
                        # form.submit() on the button's parent form if nothing
                        # navigated within the timeout.
                        nav_timeout = min(timeout, 30000)
                        navigated = False
                        try:
                            async with page.expect_navigation(timeout=nav_timeout, wait_until="load"):
                                await page.click(selector, timeout=timeout)
                            navigated = True
                        except Exception:
                            navigated = False
                        if not navigated:
                            # Click might have succeeded but no navigation fired.
                            # Give any pending JS (LeadId / TrustedForm token
                            # collectors that attach onsubmit handlers and only
                            # populate hidden fields on the first click) a brief
                            # window to finish, then call plain form.submit()
                            # which BYPASSES onsubmit handlers and forces the
                            # POST through.
                            try:
                                await page.wait_for_timeout(2500)
                            except Exception:
                                pass
                            try:
                                async with page.expect_navigation(timeout=nav_timeout, wait_until="load"):
                                    await page.evaluate(
                                        "(sel) => {"
                                        "  var el = document.querySelector(sel);"
                                        "  var f = el && (el.form || el.closest('form'));"
                                        "  if (f) { try { f.submit(); } catch(e) {} }"
                                        "}",
                                        selector,
                                    )
                                navigated = True
                            except Exception:
                                pass
                        # Best-effort wait for the post-navigation page to
                        # settle (non-fatal if already idle).
                        try:
                            await page.wait_for_load_state("networkidle", timeout=15000)
                        except Exception:
                            pass
                    else:
                        await page.click(selector, timeout=timeout)
                elif action == "fill":
                    await page.fill(selector, str(value), timeout=timeout)
                elif action == "type":
                    # Slower per-char typing (more human)
                    await page.type(selector, str(value), delay=int(step.get("delay") or 50), timeout=timeout)
                elif action == "select":
                    await page.select_option(selector, value=str(value), timeout=timeout)
                elif action == "check":
                    await page.check(selector, timeout=timeout)
                elif action == "uncheck":
                    await page.uncheck(selector, timeout=timeout)
                elif action == "press":
                    await page.press(selector or "body", value or "Enter", timeout=timeout)
                elif action == "wait":
                    await page.wait_for_timeout(int(step.get("ms") or 1000))
                elif action == "wait_for_selector":
                    await page.wait_for_selector(selector, timeout=timeout, state=step.get("state") or "visible")
                elif action in ("wait_for_navigation", "wait_for_load", "wait_for_networkidle"):
                    try:
                        await page.wait_for_load_state("networkidle", timeout=timeout)
                    except Exception:
                        pass
                elif action == "scroll":
                    try:
                        if selector:
                            el = await page.query_selector(selector)
                            if el:
                                await el.scroll_into_view_if_needed()
                        else:
                            await page.evaluate(f"window.scrollBy(0,{int(step.get('y') or 500)})")
                    except Exception:
                        pass
                elif action == "evaluate":
                    js = _substitute(step.get("script") or step.get("js") or "", row)
                    await page.evaluate(js)
                elif action == "screenshot":
                    # User-named intermediate screenshot — skipped here; main
                    # final screenshot is captured by the job loop.
                    pass
                else:
                    if not optional:
                        return {"status": "failed", "error": f"Unknown action '{action}' at step {idx+1}", "executed_steps": executed}
                executed += 1
            except Exception as e:
                if optional:
                    executed += 1
                    continue
                # ── Self-heal: ask AI to propose a recovery action ──────
                if self_heal and heal_used < MAX_HEAL:
                    heal_used += 1
                    try:
                        heal_action = await _try_self_heal(page, step, str(e))
                    except Exception:
                        heal_action = None
                    if heal_action:
                        try:
                            await _execute_automation_steps(
                                page, row, [heal_action],
                                skip_captcha=skip_captcha, self_heal=False,
                            )
                        except Exception:
                            pass
                        # Retry the original step ONCE after recovery
                        try:
                            await _dispatch_single_action(
                                page, action, selector, value, step, timeout,
                                wait_nav, row,
                            )
                            executed += 1
                            continue
                        except Exception as e2:
                            return {"status": "failed",
                                    "error": f"Step {idx+1} ({action}) failed after self-heal: {str(e2)[:200]}",
                                    "executed_steps": executed}
                return {"status": "failed", "error": f"Step {idx+1} ({action}) failed: {str(e)[:200]}", "executed_steps": executed}
        return {"status": "ok", "executed_steps": executed}
    except Exception as e:
        return {"status": "failed", "error": f"Automation crashed: {str(e)[:200]}", "executed_steps": executed}


async def _dispatch_single_action(page: Page, action: str, selector: str,
                                  value: Any, step: Dict[str, Any],
                                  timeout: int, wait_nav: bool,
                                  row: Dict[str, Any]) -> None:
    """Run one action (no retry, no self-heal). Raises on failure."""
    if action == "goto":
        await page.goto(value or selector, timeout=timeout, wait_until="domcontentloaded")
    elif action == "click":
        await page.click(selector, timeout=timeout)
        if wait_nav:
            try:
                await page.wait_for_load_state("networkidle", timeout=20000)
            except Exception:
                pass
    elif action == "fill":
        await page.fill(selector, str(value), timeout=timeout)
    elif action == "type":
        await page.type(selector, str(value), delay=int(step.get("delay") or 50), timeout=timeout)
    elif action == "select":
        await page.select_option(selector, value=str(value), timeout=timeout)
    elif action == "check":
        await page.check(selector, timeout=timeout)
    elif action == "uncheck":
        await page.uncheck(selector, timeout=timeout)
    elif action == "press":
        await page.press(selector or "body", value or "Enter", timeout=timeout)
    elif action == "wait":
        await page.wait_for_timeout(int(step.get("ms") or 1000))
    elif action == "wait_for_selector":
        await page.wait_for_selector(selector, timeout=timeout, state=step.get("state") or "visible")
    elif action in ("wait_for_navigation", "wait_for_load", "wait_for_networkidle"):
        await page.wait_for_load_state("networkidle", timeout=timeout)
    elif action == "scroll":
        if selector:
            el = await page.query_selector(selector)
            if el:
                await el.scroll_into_view_if_needed()
        else:
            await page.evaluate(f"window.scrollBy(0,{int(step.get('y') or 500)})")
    elif action == "evaluate":
        js = _substitute(step.get("script") or step.get("js") or "", row)
        await page.evaluate(js)


async def _try_self_heal(page: Page, failed_step: Dict[str, Any],
                         error_msg: str) -> Optional[Dict[str, Any]]:
    """Take a screenshot + ask Gemini for a recovery action. Returns a step
    dict or None. Keeps the call short so it doesn't stall the job."""
    try:
        from ai_automation_generator import suggest_self_heal_action
    except Exception as e:
        logger.warning(f"self-heal disabled (import failed): {e}")
        return None

    try:
        import tempfile, os as _os
        tmpdir = tempfile.gettempdir()
        path = _os.path.join(tmpdir, f"rut_heal_{uuid.uuid4().hex[:8]}.png")
        try:
            await page.screenshot(path=path, full_page=False, timeout=5000)
        except Exception:
            return None
        title = ""
        url = ""
        try:
            title = await page.title()
            url = page.url
        except Exception:
            pass
        action = await suggest_self_heal_action(
            screenshot_path=path,
            page_title=title,
            page_url=url,
            failed_step={**failed_step, "_error": error_msg[:200]},
        )
        try:
            _os.remove(path)
        except Exception:
            pass
        return action
    except Exception as e:
        logger.warning(f"self-heal call failed: {e}")
        return None


async def _multi_step_fill(page: Page, row: Dict[str, Any]) -> Dict[str, Any]:
    max_steps = 6
    for step in range(max_steps):
        await page.wait_for_timeout(500 + random.randint(0, 400))
        await _dismiss_popups(page)
        fill_info = await _fill_form(page, row)
        step_filled = len(fill_info.get("filled") or [])
        if step == 0 and step_filled == 0:
            # Diagnostic: log page URL + visible input count so we can see why
            try:
                cur_url = page.url
                vis_inputs = await page.evaluate(
                    "()=>Array.from(document.querySelectorAll('input,select,textarea'))"
                    ".filter(e=>e.type!=='hidden'&&e.offsetParent).length"
                )
                err = f"No fillable fields matched (url={cur_url[:120]} visible_inputs={vis_inputs})"
            except Exception:
                err = "No fillable fields matched"
            return {"status": "no_fields_matched", "error": err}
        if step > 0 and step_filled == 0:
            return {"status": "ok"}
        await _tick_consent_checkboxes(page)
        try:
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight/2)")
        except Exception:
            pass
        start_url = page.url
        await _click_submit(page)
        for _ in range(2):
            try:
                await page.wait_for_load_state("networkidle", timeout=8000)
            except Exception:
                pass
            await page.wait_for_timeout(900)
            if page.url != start_url:
                break
            if await _dismiss_review_modal(page):
                await _tick_consent_checkboxes(page)
                await page.wait_for_timeout(400)
                await _click_submit(page)
                continue
            break
        if page.url == start_url:
            return {"status": "submitted_but_no_redirect"}
    return {"status": "ok"}


async def _log_click_for_link(entry: Dict[str, Any], job_info: Dict[str, Any], main_db):
    """Mirror the /api/t/ tracker: create a click document in the link
    owner's per-user DB and bump the link's click counter. Called after
    every RUT visit so the dashboard Clicks page + duplicate-IP detection
    work even when the browser hits the offer URL directly (auto-swapped
    from a localhost tracker URL)."""
    import uuid as _uuid
    link_id = job_info.get("link_id")
    owner_id = job_info.get("link_owner_id")
    short_code = job_info.get("link_short_code") or ""
    if not link_id or not owner_id or main_db is None:
        return
    try:
        # Access the per-user DB on the same client. IMPORTANT: Must match
        # server.py::get_user_db() exactly — that helper uses a 20-char
        # truncated, underscore-normalised key:
        #     f"trackmaster_user_{user_id.replace('-', '_')[:20]}"
        # If we use the raw owner_id (with hyphens) here, the click docs go
        # into a SEPARATE database and the dashboard / Clicks page reads
        # from the truncated DB and sees ZERO clicks — exactly the bug
        # users have reported ("tracker link use kia pr click count nahi hoa").
        client = main_db.client
        db_name = f"trackmaster_user_{owner_id.replace('-', '_')[:20]}"
        user_db = client[db_name]

        exit_ip = (entry.get("exit_ip") or "").strip()
        is_vpn = bool(entry.get("status") == "skipped_vpn" or entry.get("is_vpn"))
        ua = entry.get("ua") or ""
        device_display = entry.get("device_name") or entry.get("os") or "Unknown"

        click_doc = {
            "id": str(_uuid.uuid4()),
            "click_id": str(_uuid.uuid4()),
            "link_id": link_id,
            "user_id": owner_id,
            "short_code": short_code,
            "ip_address": exit_ip or "unknown",
            "ipv4": exit_ip if exit_ip and ":" not in exit_ip else "",
            "all_ips": [exit_ip] if exit_ip else [],
            "country": entry.get("country") or "Unknown",
            "city": entry.get("city") or "",
            "timezone": entry.get("timezone") or "",
            "is_vpn": is_vpn,
            "is_proxy": bool(entry.get("proxy")),
            "vpn_score": 0,
            "user_agent": ua,
            "device": (entry.get("os") or "desktop").lower(),
            "device_type": (entry.get("os") or "desktop").lower(),
            "device_display": device_display,
            "device_brand": "",
            "device_model": "",
            "os_name": entry.get("os") or "",
            "os_version": "",
            "browser": "Chrome",
            "browser_version": "",
            "referrer": "",
            "referrer_source": "rut",
            "referrer_source_name": "Real User Traffic",
            "source": "real_user_traffic",
            "visit_status": entry.get("status") or "",
            "final_url": entry.get("final_url") or "",
            "conversion_page_reached": bool(entry.get("conversion_page_reached")),
            "created_at": entry.get("timestamp") or datetime.now(timezone.utc).isoformat(),
        }
        await user_db.clicks.insert_one(click_doc)
        # Bump link-level click counter on the main DB
        await main_db.links.update_one({"id": link_id}, {"$inc": {"clicks": 1}})
    except Exception as e:
        # Best-effort — never crash the visit because click logging failed
        logger.warning(f"RUT click log failed (job_id-unknown): {e}")


async def _record(
    job_id: str,
    entry: Dict[str, Any],
    report: List[Dict[str, Any]],
    lock: asyncio.Lock,
    db,
):
    async with lock:
        j = RUT_JOBS.setdefault(job_id, {})
        j["processed"] = int(j.get("processed") or 0) + 1
        s = entry.get("status", "failed")
        key_map = {
            "ok": "succeeded",
            "skipped_captcha": "skipped_captcha",
            "skipped_country": "skipped_country",
            "skipped_os": "skipped_os",
            "skipped_duplicate_ip": "skipped_duplicate_ip",
            "skipped_vpn": "skipped_vpn",
            "skipped_state_mismatch": "skipped_state_mismatch",
            "invalid_data": "invalid_data",
        }
        counter_key = key_map.get(s, "failed")
        j[counter_key] = int(j.get(counter_key) or 0) + 1
        # Conversion counter: visits where final URL redirected OFF the form page
        if entry.get("conversion_page_reached"):
            j["conversions"] = int(j.get("conversions") or 0) + 1
        report.append(entry)

        # ── Log this visit as a click against the link (so dashboard's
        #    Clicks page + duplicate-IP detection both see RUT traffic) ──
        try:
            await _log_click_for_link(entry, j, db)
        except Exception:
            pass

        events = j.setdefault("events", [])
        events.append({
            "row": entry["visit_index"],
            "status": entry["status"],
            "proxy": entry["proxy"],
            "exit_ip": entry["exit_ip"],
            "country": entry["country"],
            "city": entry["city"],
            "device": f"{entry.get('device_name') or entry['os']} · {entry['viewport']}",
            "final_url": entry["final_url"],
            "conversion_page_reached": bool(entry.get("conversion_page_reached")),
            "error": (entry["error"] or "")[:140],
            "ts": entry["timestamp"],
        })
        if len(events) > 80:
            del events[:-80]
        if db is not None and j["processed"] % 5 == 0:
            try:
                await _persist(db, job_id)
            except Exception:
                pass


def _did_reach_conversion(landing_url: str, final_url: str) -> bool:
    """True if after form submit the user ended up on a DIFFERENT host than
    the original landing (e.g. thnkspg.com after stimulusassistforall.com),
    OR on the same host but on a different page stem than the form page
    (indexform / index-form / index). Indicates the offer accepted the lead
    and redirected the user forward."""
    try:
        from urllib.parse import urlparse
        lu = urlparse(landing_url or "")
        fu = urlparse(final_url or "")
        lh = (lu.netloc or "").lower().lstrip("www.")
        fh = (fu.netloc or "").lower().lstrip("www.")
        if not fh:
            return False
        if lh and fh and lh != fh:
            # Different second-level domain → classic conversion redirect
            lh_root = ".".join(lh.split(".")[-2:])
            fh_root = ".".join(fh.split(".")[-2:])
            if lh_root != fh_root:
                return True
        # Same host: compare path stems. Landing is usually index-form.php
        # or indexform.php; a conversion moves to offers-flow.php etc.
        form_stems = ("index-form", "indexform", "/index.php", "/index-form.php")
        lp = (lu.path or "").lower()
        fp = (fu.path or "").lower()
        if fp and lp != fp:
            if not any(s in fp for s in form_stems):
                return True
        return False
    except Exception:
        return False


# ─── Strict thank-you page detection ──────────────────────────────
# Per user request: only count a visit as a CONVERSION (and only take the
# final screenshot) when we are confident the browser reached the offer's
# thank-you / confirmation page. We combine three signals:
#   1. Host root changed vs. the landing page (e.g. stimulusassistforall.com
#      → thnkspg.com)
#   2. URL path / query contains a thank-you / success / claim / offer keyword
#   3. Page title or body text contains strong thank-you text (e.g.
#      "Claim Your $750", "Thank You", "Ways to Earn", "Congratulations")
# At least TWO of the three positive signals must match → avoids counting
# captcha-redirects, error pages, or same-host follow-ups as conversions.
_THANKYOU_URL_KEYWORDS = [
    "thank", "thanks", "thnk", "/ty", "ty.php", "thnks", "thnkspg",
    "success", "confirm", "confirmation", "completed",
    "claim", "offer", "offers-flow", "offer-flow", "offerwall", "offer-wall",
    "reward", "congrat", "congrats", "received", "submitted",
]
_THANKYOU_TEXT_KEYWORDS = [
    "thank you", "thank-you", "thankyou", "thanks for",
    "congratulations", "congrats",
    "claim your", "claim $", "claim 1 deal", "claim 1 prize",
    "your reward", "your prize", "your claim",
    "successfully submitted", "submission received", "submission successful",
    "we received", "we've received", "order confirmation",
    "complete paid offers", "complete the offers", "ways to earn",
    "pending offers", "pending offer", "offer wall",
    "you qualify", "you're qualified", "you have been matched",
]
# Page texts that STRONGLY indicate we're still on the form page (not converted)
_FORM_PAGE_TEXT_NEGATIVES = [
    "enter your first name", "fill out the form", "complete the form below",
    "please fill", "submit below", "please correct the errors",
]


def _is_thank_you_page(landing_url: str, final_url: str,
                       page_text: str = "", page_title: str = "") -> bool:
    """Strict thank-you / conversion page check. Returns True only when at
    least TWO of {host-change, URL-keyword, page-text-keyword} match."""
    try:
        from urllib.parse import urlparse
        lu = urlparse(landing_url or "")
        fu = urlparse(final_url or "")
    except Exception:
        return False

    if not fu.netloc:
        return False

    lh = (lu.netloc or "").lower().lstrip("www.")
    fh = (fu.netloc or "").lower().lstrip("www.")

    # 1. host root change
    host_changed = False
    if lh and fh and lh != fh:
        lh_root = ".".join(lh.split(".")[-2:])
        fh_root = ".".join(fh.split(".")[-2:])
        host_changed = lh_root != fh_root

    # 2. URL keyword match
    full_url = ((fu.geturl() or "") + " " + (fu.path or "") + " " + (fu.query or "")).lower()
    url_keyword_hit = any(k in full_url for k in _THANKYOU_URL_KEYWORDS)

    # 3. page text / title keyword match
    text_combined = ((page_title or "") + " " + (page_text or "")).lower()[:6000]
    text_keyword_hit = any(k in text_combined for k in _THANKYOU_TEXT_KEYWORDS)

    # Strong negative: if we see clear form-page text, require EXTRA evidence
    is_still_on_form = any(k in text_combined for k in _FORM_PAGE_TEXT_NEGATIVES)

    positives = sum([host_changed, url_keyword_hit, text_keyword_hit])
    if is_still_on_form:
        # Need all three signals to overrule strong form-page evidence
        return positives >= 3
    return positives >= 2


def _write_excel_report(out_path: Path, report: List[Dict[str, Any]]):
    if not report:
        df = pd.DataFrame([{"info": "no visits completed"}])
    else:
        df = pd.DataFrame(report)
        # nice column order
        preferred = [
            "visit_index", "status", "conversion_page_reached", "proxy", "exit_ip",
            "country", "city",
            "timezone", "locale", "os", "viewport", "device_scale_factor",
            "hardware_concurrency", "device_memory", "webgl_renderer", "canvas_seed",
            "ua", "http_status", "landing_url", "final_url", "trusted_form", "lead_id",
            "screenshot", "error", "timestamp",
        ]
        cols = [c for c in preferred if c in df.columns] + [c for c in df.columns if c not in preferred]
        df = df[cols]
    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        df.to_excel(xl, sheet_name="Visits", index=False)
        if report:
            # Summary sheet
            status_counts = df["status"].value_counts().reset_index()
            status_counts.columns = ["status", "count"]
            status_counts.to_excel(xl, sheet_name="Summary", index=False)
            # Conversion summary
            if "conversion_page_reached" in df.columns:
                conv_true = int(df["conversion_page_reached"].fillna(False).astype(bool).sum())
                conv_total = int(len(df))
                conv_ok = int(((df["status"] == "ok") & (df["conversion_page_reached"].fillna(False).astype(bool))).sum())
                conv_df = pd.DataFrame([
                    {"metric": "total_visits", "count": conv_total},
                    {"metric": "status_ok", "count": int((df["status"] == "ok").sum())},
                    {"metric": "conversion_page_reached", "count": conv_true},
                    {"metric": "conversion_and_ok", "count": conv_ok},
                    {"metric": "conversion_rate_pct",
                     "count": round(100.0 * conv_true / conv_total, 2) if conv_total else 0.0},
                ])
                conv_df.to_excel(xl, sheet_name="Conversions", index=False)


def _finalise(job_id: str, status: str, error: str = ""):
    j = RUT_JOBS.setdefault(job_id, {})
    j["status"] = status
    if error:
        j["error"] = error
    j["finished_at"] = datetime.now(timezone.utc).isoformat()


async def _finalise_and_persist(db, job_id: str, status: str, error: str = ""):
    """Same as _finalise but ALSO writes the failed/stopped state to MongoDB
    so the Past Jobs row + REST endpoint reflect the error message instead
    of leaving the job stuck on 'queued' forever."""
    _finalise(job_id, status, error)
    if db is not None:
        try:
            await _persist(db, job_id)
        except Exception as e:
            logger.debug(f"_finalise_and_persist persist failed: {e}")


async def _persist(db, job_id: str):
    j = RUT_JOBS.get(job_id, {})
    if not j:
        return
    # Filter out non-serializable entries (e.g. asyncio.Event)
    safe = {k: v for k, v in j.items() if not k.startswith("_")}
    try:
        await db.real_user_traffic_jobs.update_one(
            {"job_id": job_id},
            {"$set": {**safe, "job_id": job_id}},
            upsert=True,
        )
    except Exception as e:
        logger.debug(f"persist real_user_traffic_jobs failed: {e}")


def request_job_cancel(job_id: str) -> bool:
    """Flip the in-memory cancel flag on a running job. Returns True if the
    job was found and signalled, False otherwise (job finished / unknown)."""
    j = RUT_JOBS.get(job_id)
    if not j:
        return False
    ev = j.get("_cancel_event")
    if ev is None:
        return False
    try:
        ev.set()
        j["cancel_requested_at"] = datetime.now(timezone.utc).isoformat()
        return True
    except Exception:
        return False


def create_rut_job(
    job_id: str,
    user_id: str,
    target_url: str,
    total: int,
    form_fill_enabled: bool,
) -> Dict[str, Any]:
    RUT_JOBS[job_id] = {
        "job_id": job_id,
        "user_id": user_id,
        "target_url": target_url,
        "total": total,
        "form_fill_enabled": form_fill_enabled,
        "status": "queued",
        "processed": 0,
        "succeeded": 0,
        "skipped_captcha": 0,
        "skipped_country": 0,
        "skipped_os": 0,
        "skipped_duplicate_ip": 0,
        "skipped_vpn": 0,
        "failed": 0,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    return RUT_JOBS[job_id]


def cleanup_rut_job(job_id: str):
    import shutil
    d = RESULTS_ROOT / job_id
    if d.exists():
        shutil.rmtree(d, ignore_errors=True)
    RUT_JOBS.pop(job_id, None)


# ── Live step-log (for the "what's happening now" modal) ────────────
# Each job keeps a bounded ring-buffer of recent steps so the UI can stream
# them without any backend cost until the modal is actually opened.
_MAX_LIVE_STEPS = 300

def push_live_step(job_id: str, visit: int, stage: str, status: str, detail: str = "",
                   screenshot: str = ""):
    j = RUT_JOBS.get(job_id)
    if j is None:
        return
    buf = j.setdefault("live_steps", [])
    buf.append({
        "idx": len(buf) + 1,
        "visit": visit,
        "stage": stage,              # "setup" | "geo" | "filter" | "browser" | "form" | "submit" | "done"
        "status": status,            # "info" | "ok" | "skipped" | "failed"
        "detail": (detail or "")[:200],
        "screenshot": screenshot,    # filename only; served by /jobs/{id}/screenshot/{file}
        "ts": datetime.now(timezone.utc).isoformat(),
    })
    # Cap buffer
    if len(buf) > _MAX_LIVE_STEPS:
        del buf[:-_MAX_LIVE_STEPS]


def get_live_steps(job_id: str, since: int = 0) -> Dict[str, Any]:
    """Return steps with idx > since (used by the frontend modal)."""
    j = RUT_JOBS.get(job_id)
    if j is None:
        return {"steps": [], "cursor": since, "running": False}
    buf = j.get("live_steps") or []
    new = [s for s in buf if s.get("idx", 0) > since]
    return {
        "steps": new,
        "cursor": buf[-1]["idx"] if buf else since,
        "running": j.get("status") in ("running", "queued"),
        "status": j.get("status", "unknown"),
        "processed": j.get("processed", 0),
        "total": j.get("total", 0),
    }
